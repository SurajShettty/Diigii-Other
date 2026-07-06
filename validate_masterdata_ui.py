"""
Masterdata validator with drag-and-drop GUI for the BSACIST multi-tab Excel template.

Validates department / programme / course / faculty / admin / student sheets
against a live tenant database and writes the following into an output folder:

  - masterdata_validated.xlsx  (original workbook + "Remarks" column)
  - student_creation.xlsx      (OK student rows in bulk-upload format)
  - staff_creation.xlsx        (OK faculty + admin rows in bulk-upload format)

Usage:
    python validate_masterdata_ui.py

Launches a desktop window where you can drag-and-drop the Excel file,
pick the tenant schema, select which tabs to validate, and choose the
output folder.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd

# mysql-connector-python may not be installed in every venv; guard the import.
try:
    import mysql.connector
except ImportError as exc:  # pragma: no cover
    mysql = None
    _mysql_import_error = exc
else:
    _mysql_import_error = None

from db_env import get_db_config, list_schemas

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEET_MAP = {
    "department": {
        "sheet": "Department Data",
        "header_row": 0,
    },
    "programme": {
        "sheet": "Programme Degree Data",
        "header_row": 1,  # row 1 in Excel is a title; headers are on row 2
    },
    "course": {
        "sheet": "Course Data",
        "header_row": 0,
    },
    "faculty": {
        "sheet": "Faculty Data",
        "header_row": 0,
    },
    "admin": {
        "sheet": "Administrative Staff Data",
        "header_row": 0,
    },
    "student": {
        "sheet": "Students Data",
        "header_row": 0,
    },
}

REMARKS_COL = "Remarks"
OK_TEXT = "OK"

EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")
PHONE_RE = re.compile(r"^\d{10}$")


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def normalise_whitespace(value: Any) -> str:
    """Return a stripped string, treating null-like values as empty."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none", "nat", "null"}:
        return ""
    return s


def is_blank(value: Any) -> bool:
    return normalise_whitespace(value) == ""


def clean_phone(value: Any) -> Optional[str]:
    """Return 10-digit phone string or None if invalid/empty."""
    if is_blank(value):
        return None
    s = normalise_whitespace(value)
    # Excel sometimes stores phone numbers as floats ending with .0
    if s.endswith(".0"):
        s = s[:-2]
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits if len(digits) == 10 else None


def is_valid_email(value: Any) -> bool:
    s = normalise_whitespace(value)
    return bool(s and EMAIL_RE.fullmatch(s))


def normalise_gender(value: Any) -> Optional[str]:
    s = normalise_whitespace(value).lower()
    if s in {"male", "m"}:
        return "Male"
    if s in {"female", "f"}:
        return "Female"
    if s in {"other", "o"}:
        return "Other"
    return None


def to_int(value: Any) -> Optional[int]:
    s = normalise_whitespace(value)
    if not s:
        return None
    try:
        f = float(s)
        if f.is_integer():
            return int(f)
    except ValueError:
        pass
    return None


def _intake_key(value: str) -> str:
    """Normalise an intake string for lenient comparison (ignore case/separators)."""
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _format_intake(raw_intake: str, programme: str, batch_year: Optional[int]) -> str:
    """Return the standard intake format; case of raw value is ignored."""
    if not programme or batch_year is None:
        return raw_intake
    return f"{programme}-{batch_year}-intake"


def find_column(df: pd.DataFrame, prefixes: List[str]) -> Optional[str]:
    """Find the first column whose header starts with one of the prefixes (case-insensitive)."""
    prefixes = [p.strip().lower() for p in prefixes]
    for col in df.columns:
        col_norm = str(col).strip().lower()
        for prefix in prefixes:
            if col_norm.startswith(prefix):
                return col
    return None


def load_sheet(path: str, sheet: str, header_row: int = 0) -> pd.DataFrame:
    """Load one sheet, normalise blanks, return dtype=str DataFrame.

    Adds an internal `__excel_row__` column so output remarks can be written
    back to the correct Excel row even if blank rows are skipped.
    """
    df = pd.read_excel(path, sheet_name=sheet, header=header_row, dtype=str)
    # Strip column names
    df.columns = [str(c).strip() for c in df.columns]
    # Replace common null-like strings and pandas NaNs with empty string
    df = df.replace(to_replace=r"(?i)^\s*(nan|none|nat|null)\s*$", value="", regex=True)
    df = df.fillna("")
    # Excel row number for the first data row is header_row + 2 (1-based)
    df["__excel_row__"] = df.index + header_row + 2
    return df


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------


class DbCache:
    """Lightweight read-only cache for the tenant tables used by validators."""

    def __init__(self, conn):
        self.conn = conn
        self._cache: Dict[str, List[Dict[str, Any]]] = {}

    def fetch(self, table: str) -> List[Dict[str, Any]]:
        if table not in self._cache:
            cur = self.conn.cursor(dictionary=True)
            cur.execute(f"SELECT * FROM {table}")
            rows = cur.fetchall()
            cur.close()
            # JSON-safe stringification for dates/decimals/bytes
            for row in rows:
                for k, v in row.items():
                    if v is not None and not isinstance(v, (str, int, float, bool)):
                        row[k] = str(v)
            self._cache[table] = rows
        return self._cache[table]

    def column_names(self, table: str) -> List[str]:
        rows = self.fetch(table)
        if rows:
            return list(rows[0].keys())
        # Table is empty; get column names from the cursor description.
        cur = self.conn.cursor()
        try:
            cur.execute(f"SELECT * FROM {table} LIMIT 0")
            cols = [d[0] for d in cur.description] if cur.description else []
            cur.fetchall()  # consume any residual result set
        finally:
            cur.close()
        return cols

    def find_col(self, table: str, candidates: List[str]) -> Optional[str]:
        """Return the first available column name matching one of the candidates."""
        cols = {c.lower() for c in self.column_names(table)}
        for cand in candidates:
            if cand.lower() in cols:
                return cand
        # Fuzzy: candidate substring in column name
        for col in self.column_names(table):
            col_l = col.lower()
            for cand in candidates:
                if cand.lower() in col_l:
                    return col
        return None

    def value_set(self, table: str, column: str, lower: bool = False) -> Set[str]:
        rows = self.fetch(table)
        values = set()
        for row in rows:
            v = row.get(column)
            if v is not None and str(v).strip():
                values.add(str(v).strip().lower() if lower else str(v).strip())
        return values

    def department_names(self) -> Set[str]:
        col = self.find_col("department", ["name", "department_name", "dept_name"])
        if not col:
            raise ValueError(
                f"Could not find a name column in department. Available columns: {self.column_names('department')}"
            )
        return self.value_set("department", col, lower=True)

    def programme_names(self) -> Set[str]:
        col = self.find_col("programme", ["name", "programme_name", "program_name", "programme"])
        if not col:
            raise ValueError(
                f"Could not find a name column in programme. Available columns: {self.column_names('programme')}"
            )
        return self.value_set("programme", col, lower=True)

    def authenticator_emails(self) -> Set[str]:
        col = self.find_col("authenticator", ["email", "email_id", "mail"])
        if not col:
            return set()
        return self.value_set("authenticator", col, lower=True)

    def authenticator_phones(self) -> Set[str]:
        col = self.find_col("authenticator", ["phone_number", "mobile_number", "phone", "mobile"])
        if not col:
            return set()
        phones = set()
        for row in self.fetch("authenticator"):
            v = row.get(col)
            if v is not None:
                digits = "".join(ch for ch in str(v) if ch.isdigit())
                if len(digits) == 10:
                    phones.add(digits)
        return phones

    def user_attribute_registration_ids(self) -> Set[str]:
        """Return existing registration ids from user_attributes table."""
        col = self.find_col(
            "user_attributes",
            ["registration_id", "registration_no", "roll_number", "student_id", "enrollment_no", "enrollment_id"],
        )
        if not col:
            return set()
        return self.value_set("user_attributes", col, lower=True)

    def quota_names(self) -> Set[str]:
        col = self.find_col("quota", ["name", "quota_name"])
        if not col:
            return set()
        return self.value_set("quota", col, lower=True)

    def intake_names(self) -> Set[str]:
        col = self.find_col(
            "programme_batch_intake",
            ["name", "intake_name", "batch_name", "programme_batch_intake_name"],
        )
        if not col:
            return set()
        return self.value_set("programme_batch_intake", col, lower=True)

    def intakes_mapped_to_terms(self) -> Set[str]:
        """Return intake identifiers that appear in term_programme_batch."""
        if not self.fetch("term_programme_batch"):
            return set()

        # Try to identify the column that points to programme_batch_intake
        intake_col = self.find_col(
            "term_programme_batch",
            ["programme_batch_intake_id", "intake_id", "programme_batch_intake", "intake"],
        )
        if not intake_col:
            # Fallback: try to match values against known intake names
            intake_names = self.intake_names()
            tpb_cols = self.column_names("term_programme_batch")
            for col in tpb_cols:
                vals = self.value_set("term_programme_batch", col, lower=True)
                if vals & intake_names:
                    intake_col = col
                    break

        if not intake_col:
            return set()

        mapped = self.value_set("term_programme_batch", intake_col, lower=True)

        # Also try to resolve IDs to intake names
        intake_id_col = self.find_col("programme_batch_intake", ["id"])
        intake_name_col = self.find_col("programme_batch_intake", ["name", "intake_name", "batch_name"])
        if intake_id_col and intake_name_col:
            id_to_name = {}
            for row in self.fetch("programme_batch_intake"):
                pk = row.get(intake_id_col)
                name = row.get(intake_name_col)
                if pk is not None and name is not None:
                    id_to_name[str(pk).strip().lower()] = str(name).strip().lower()
            extra = {id_to_name.get(v) for v in mapped if v in id_to_name}
            mapped.update({x for x in extra if x})

        return mapped

    def intake_name_keys(self) -> Set[str]:
        """Return normalised intake-name keys for lenient matching."""
        return {_intake_key(name) for name in self.intake_names()}

    def mapped_intake_keys(self) -> Set[str]:
        """Return normalised mapped-intake keys for lenient matching."""
        return {_intake_key(name) for name in self.intakes_mapped_to_terms()}

    def department_layer_map(self) -> Dict[str, str]:
        """Return {lowercase_department_name: layer_name} using the user's query logic."""
        mapping: Dict[str, str] = {}

        # Ensure tables are cached
        dept_rows = self.fetch("department")
        entity_rows = self.fetch("institution_entity")
        layer_rows = self.fetch("institution_layer")

        if not dept_rows:
            return mapping

        # Identify columns
        dept_name_col = self.find_col("department", ["name", "department_name", "dept_name"])
        parent_entity_col = self.find_col("department", ["parent_entity_id"])
        entity_id_col = self.find_col("institution_entity", ["id"])
        entity_name_col = self.find_col("institution_entity", ["name"])
        layer_id_col = self.find_col("institution_layer", ["id"])
        layer_name_col = self.find_col("institution_layer", ["name"])

        if not all([dept_name_col, parent_entity_col, entity_id_col, layer_id_col, layer_name_col]):
            return mapping

        # entity_id -> layer_name
        entity_to_layer: Dict[str, str] = {}
        if layer_rows and entity_rows:
            layer_id_to_name = {
                str(row.get(layer_id_col)).strip().lower(): str(row.get(layer_name_col)).strip()
                for row in layer_rows
                if row.get(layer_id_col) is not None and row.get(layer_name_col) is not None
            }
            for row in entity_rows:
                eid = str(row.get(entity_id_col)).strip().lower() if row.get(entity_id_col) is not None else None
                lid = str(row.get(layer_id_col)).strip().lower() if row.get(layer_id_col) is not None else None
                if eid and lid and lid in layer_id_to_name:
                    entity_to_layer[eid] = layer_id_to_name[lid]

        # department_name -> layer_name
        for row in dept_rows:
            dept_name = normalise_whitespace(row.get(dept_name_col, ""))
            parent_id = str(row.get(parent_entity_col)).strip().lower() if row.get(parent_entity_col) is not None else ""
            if dept_name and parent_id and parent_id in entity_to_layer:
                mapping[dept_name.lower()] = entity_to_layer[parent_id]

        return mapping


class FreshContext:
    """Validation context for fresh instances (no DB data yet).

    Provides the same lookup interface as DbCache but sources department and
    programme references from the masterdata tabs instead of the database.
    """

    def __init__(self, assessed_sheets: Dict[str, pd.DataFrame]):
        self._assessed = assessed_sheets
        self._dept_names = self._build_dept_names()
        self._prog_names = self._build_prog_names()
        self._prog_duration_map = self._build_prog_duration_map()

    def _build_dept_names(self) -> Set[str]:
        df = self._assessed.get("department")
        if df is None:
            return set()
        col = find_column(df, ["name"])
        if not col:
            return set()
        return {normalise_whitespace(v).lower() for v in df[col] if not is_blank(v)}

    def _build_prog_names(self) -> Set[str]:
        df = self._assessed.get("programme")
        if df is None:
            return set()
        col = find_column(df, ["programme name"])
        if not col:
            return set()
        return {normalise_whitespace(v).lower() for v in df[col] if not is_blank(v)}

    def _build_prog_duration_map(self) -> Dict[str, int]:
        df = self._assessed.get("programme")
        duration_map: Dict[str, int] = {}
        if df is None:
            return duration_map
        name_col = find_column(df, ["programme name"])
        dur_col = find_column(df, ["duration"])
        if not name_col or not dur_col:
            return duration_map
        for _, row in df.iterrows():
            name = normalise_whitespace(row.get(name_col, ""))
            dur = to_int(row.get(dur_col, ""))
            if name and dur is not None:
                duration_map[name.lower()] = dur
        return duration_map

    def department_names(self) -> Set[str]:
        return self._dept_names

    def programme_names(self) -> Set[str]:
        return self._prog_names

    def programme_duration(self, programme_name: str) -> Optional[int]:
        return self._prog_duration_map.get(programme_name.lower())

    def quota_names(self) -> Set[str]:
        return set()

    def intake_name_keys(self) -> Set[str]:
        return set()

    def mapped_intake_keys(self) -> Set[str]:
        return set()

    def authenticator_emails(self) -> Set[str]:
        return set()

    def authenticator_phones(self) -> Set[str]:
        return set()

    def user_attribute_registration_ids(self) -> Set[str]:
        return set()

    def department_layer_map(self) -> Dict[str, str]:
        return {}


# ---------------------------------------------------------------------------
# Validators
# ---------------------------------------------------------------------------


def _is_blank_row(row: pd.Series) -> bool:
    """True when all original data columns are blank (excludes internal helpers and Remarks)."""
    cols = [
        c for c in row.index
        if not str(c).startswith("__") and str(c) != REMARKS_COL
    ]
    return all(is_blank(row[c]) for c in cols)


def validate_department(df: pd.DataFrame, db: DbCache, fresh_instance: bool = False) -> pd.DataFrame:
    dept_names = db.department_names()
    remarks = []

    name_col = find_column(df, ["name"])
    code_col = find_column(df, ["code"])
    type_col = find_column(df, ["type"])

    for _, row in df.iterrows():
        if _is_blank_row(row):
            remarks.append("")
            continue
        msgs = []

        if not name_col:
            msgs.append("Mandatory field 'Name' is missing in masterdata")
        else:
            name = normalise_whitespace(row[name_col])
            if not name:
                msgs.append("Name is mandatory")
            elif not fresh_instance and name.lower() in dept_names:
                msgs.append("Department already exists in system")

        if code_col:
            if is_blank(row[code_col]):
                msgs.append("Code is mandatory")
        else:
            msgs.append("Mandatory field 'Code' is missing in masterdata")

        if type_col:
            if is_blank(row[type_col]):
                msgs.append("Type is mandatory")
        else:
            msgs.append("Mandatory field 'Type' is missing in masterdata")

        remarks.append("; ".join(msgs) if msgs else OK_TEXT)

    df[REMARKS_COL] = remarks
    return df


def validate_programme(df: pd.DataFrame, db: DbCache, fresh_instance: bool = False) -> pd.DataFrame:
    dept_names = db.department_names()
    prog_names = db.programme_names()
    remarks = []

    prog_col = find_column(df, ["programme name"])
    dept_col = find_column(df, ["department"])
    start_year_col = find_column(df, ["programme start year"])
    code_col = find_column(df, ["code"])
    duration_col = find_column(df, ["duration"])
    system_col = find_column(df, ["system"])

    for _, row in df.iterrows():
        if _is_blank_row(row):
            remarks.append("")
            continue
        msgs = []

        name = normalise_whitespace(row.get(prog_col, "")) if prog_col else ""
        if not name:
            msgs.append("Programme Name is mandatory")
        elif not fresh_instance and name.lower() in prog_names:
            msgs.append("Programme already exists in system")

        dept = normalise_whitespace(row.get(dept_col, "")) if dept_col else ""
        if not dept:
            msgs.append("Department is mandatory")
        elif dept.lower() not in dept_names:
            msgs.append("Department does not exist in system")

        if start_year_col:
            if is_blank(row[start_year_col]):
                msgs.append("Programme Start Year is mandatory")
            elif to_int(row[start_year_col]) is None:
                msgs.append("Programme Start Year must be a number")
        else:
            msgs.append("Mandatory field 'Programme Start Year' is missing in masterdata")

        if code_col:
            if is_blank(row[code_col]):
                msgs.append("Code is mandatory")
        else:
            msgs.append("Mandatory field 'Code' is missing in masterdata")

        duration = None
        if duration_col:
            if is_blank(row[duration_col]):
                msgs.append("Duration is mandatory")
            else:
                duration = to_int(row[duration_col])
                if duration is None:
                    msgs.append("Duration must be a number")
        else:
            msgs.append("Mandatory field 'Duration' is missing in masterdata")

        system = normalise_whitespace(row.get(system_col, "")) if system_col else ""
        if not system:
            msgs.append("System is mandatory")

        remarks.append("; ".join(msgs) if msgs else OK_TEXT)

    df[REMARKS_COL] = remarks
    return df


def validate_course(df: pd.DataFrame, db: DbCache, fresh_instance: bool = False) -> pd.DataFrame:
    dept_names = db.department_names()
    course_names = (
        set() if fresh_instance
        else db.value_set("course", db.find_col("course", ["name", "course_name"]) or "", lower=True)
    )
    course_codes = (
        set() if fresh_instance
        else db.value_set("course", db.find_col("course", ["code", "course_code"]) or "", lower=True)
    )
    remarks = []

    name_col = find_column(df, ["course name"])
    code_col = find_column(df, ["course code"])
    credits_col = find_column(df, ["total credits"])
    dept_col = find_column(df, ["department"])

    # Uploaded-record uniqueness
    uploaded_codes: Dict[str, List[int]] = {}
    for idx, row in df.iterrows():
        if code_col:
            code = normalise_whitespace(row[code_col])
            if code:
                uploaded_codes.setdefault(code.lower(), []).append(idx)

    for idx, row in df.iterrows():
        msgs = []

        name = normalise_whitespace(row.get(name_col, "")) if name_col else ""
        if not name:
            msgs.append("Course Name is mandatory")
        elif not fresh_instance and name.lower() in course_names:
            msgs.append("Course already exists in system")

        code = normalise_whitespace(row.get(code_col, "")) if code_col else ""
        if not code:
            msgs.append("Course Code is mandatory")
        else:
            if not fresh_instance and code.lower() in course_codes:
                msgs.append("Course Code already exists in system")
            if code.lower() in uploaded_codes and len(uploaded_codes[code.lower()]) > 1:
                msgs.append("Duplicate Course Code within uploaded records")

        if credits_col:
            if is_blank(row[credits_col]):
                msgs.append("Total Credits is mandatory")
            elif to_int(row[credits_col]) is None:
                try:
                    float(normalise_whitespace(row[credits_col]))
                except ValueError:
                    msgs.append("Total Credits must be numeric")
        else:
            msgs.append("Mandatory field 'Total Credits' is missing in masterdata")

        dept = normalise_whitespace(row.get(dept_col, "")) if dept_col else ""
        if not dept:
            msgs.append("Department is mandatory")
        elif dept.lower() not in dept_names:
            msgs.append("Department does not exist in system")

        remarks.append("; ".join(msgs) if msgs else OK_TEXT)

    df[REMARKS_COL] = remarks
    return df


def _validate_staff(
    df: pd.DataFrame,
    db: DbCache,
    existing_emails: Set[str],
    existing_phones: Set[str],
    uploaded_emails: Dict[str, List[int]],
    uploaded_phones: Dict[str, List[int]],
    sheet_label: str,
    fresh_instance: bool = False,
) -> pd.DataFrame:
    dept_names = db.department_names()
    auth_emails = set() if fresh_instance else db.authenticator_emails()
    auth_phones = set() if fresh_instance else db.authenticator_phones()
    existing_reg_ids = db.user_attribute_registration_ids()
    remarks = []

    emp_col = find_column(df, ["employee id"])
    name_col = find_column(df, ["name"])
    dept_col = find_column(df, ["department"])
    email_col = find_column(df, ["email"])
    phone_col = find_column(df, ["phone number"])
    gender_col = find_column(df, ["gender"])

    # Employee id uniqueness within this sheet
    emp_ids: Dict[str, List[int]] = {}
    if emp_col:
        for idx, row in df.iterrows():
            eid = normalise_whitespace(row[emp_col])
            if eid:
                emp_ids.setdefault(eid.lower(), []).append(idx)

    for idx, row in df.iterrows():
        if _is_blank_row(row):
            remarks.append("")
            continue
        msgs = []

        if emp_col:
            eid = normalise_whitespace(row[emp_col])
            if not eid:
                msgs.append("Employee Id is mandatory")
            elif eid.lower() in emp_ids and len(emp_ids[eid.lower()]) > 1:
                msgs.append("Duplicate Employee Id within uploaded records")
        else:
            msgs.append("Mandatory field 'Employee Id' is missing in masterdata")

        if name_col:
            if is_blank(row[name_col]):
                msgs.append("Name is mandatory")
        else:
            msgs.append("Mandatory field 'Name' is missing in masterdata")

        if gender_col:
            gender = normalise_gender(row[gender_col])
            if not gender:
                msgs.append("Gender must be Male, Female, or Other")
        else:
            msgs.append("Mandatory field 'Gender' is missing in masterdata")

        email = ""
        if email_col:
            email = normalise_whitespace(row[email_col])
            if not email:
                msgs.append("Email is mandatory")
            elif not is_valid_email(email):
                msgs.append("Invalid Email format")
            else:
                email_l = email.lower()
                if email_l in auth_emails:
                    msgs.append("Email already exists in system")
                elif email_l in existing_emails:
                    msgs.append("Duplicate Email across uploaded staff records")
                elif email_l in uploaded_emails and len(uploaded_emails[email_l]) > 1:
                    msgs.append("Duplicate Email within uploaded records")
        else:
            msgs.append("Mandatory field 'Email' is missing in masterdata")

        phone = None
        if phone_col:
            phone = clean_phone(row[phone_col])
            raw_phone = normalise_whitespace(row[phone_col])
            if not raw_phone:
                msgs.append("Phone Number is mandatory")
            elif phone is None:
                msgs.append("Phone Number must be 10 digits")
            else:
                if phone in auth_phones:
                    msgs.append("Phone Number already exists in system")
                elif phone in existing_phones:
                    msgs.append("Duplicate Phone Number across uploaded staff records")
                elif phone in uploaded_phones and len(uploaded_phones[phone]) > 1:
                    msgs.append("Duplicate Phone Number within uploaded records")
        else:
            msgs.append("Mandatory field 'Phone Number' is missing in masterdata")

        dept = normalise_whitespace(row.get(dept_col, "")) if dept_col else ""
        if not dept:
            msgs.append("Department is mandatory")
        elif dept.lower() not in dept_names:
            msgs.append("Department does not exist in system")

        remarks.append("; ".join(msgs) if msgs else OK_TEXT)

    df[REMARKS_COL] = remarks
    return df


def validate_faculty_and_admin(
    fac_df: pd.DataFrame,
    admin_df: pd.DataFrame,
    db: DbCache,
    fresh_instance: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    # Cross-sheet duplicates
    uploaded_emails: Dict[str, List[int]] = {}
    uploaded_phones: Dict[str, List[int]] = {}

    email_col_f = find_column(fac_df, ["email"])
    phone_col_f = find_column(fac_df, ["phone number"])
    email_col_a = find_column(admin_df, ["email"])
    phone_col_a = find_column(admin_df, ["phone number"])

    def _collect(df: pd.DataFrame, email_col, phone_col, offset: int):
        for idx, row in df.iterrows():
            if email_col:
                email = normalise_whitespace(row[email_col]).lower()
                if email:
                    uploaded_emails.setdefault(email, []).append(idx + offset)
            if phone_col:
                phone = clean_phone(row[phone_col])
                if phone:
                    uploaded_phones.setdefault(phone, []).append(idx + offset)

    _collect(fac_df, email_col_f, phone_col_f, 0)
    _collect(admin_df, email_col_a, phone_col_a, len(fac_df))

    # Existing sets from the other staff type are empty because we validate both at once
    existing_emails: Set[str] = set()
    existing_phones: Set[str] = set()

    fac_df = _validate_staff(fac_df, db, existing_emails, existing_phones, uploaded_emails, uploaded_phones, "Faculty", fresh_instance)
    admin_df = _validate_staff(admin_df, db, existing_emails, existing_phones, uploaded_emails, uploaded_phones, "Admin", fresh_instance)
    return fac_df, admin_df


def validate_student(df: pd.DataFrame, db: DbCache, fresh_instance: bool = False) -> pd.DataFrame:
    dept_names = db.department_names()
    prog_names = db.programme_names()
    quota_names = set() if fresh_instance else db.quota_names()
    intake_name_keys = set() if fresh_instance else db.intake_name_keys()
    mapped_intake_keys = set() if fresh_instance else db.mapped_intake_keys()
    auth_emails = set() if fresh_instance else db.authenticator_emails()
    auth_phones = set() if fresh_instance else db.authenticator_phones()
    existing_reg_ids = set() if fresh_instance else db.user_attribute_registration_ids()
    remarks = []

    reg_col = find_column(df, ["registration id"])
    name_col = find_column(df, ["name"])
    dept_col = find_column(df, ["department"])
    prog_col = find_column(df, ["programme", "programme/degree"])
    batch_col = find_column(df, ["batch year"])
    intake_col = find_column(df, ["intake"])
    quota_col = find_column(df, ["quota"])
    email_col = find_column(df, ["email"])
    phone_col = find_column(df, ["phone number"])
    gender_col = find_column(df, ["gender"])
    section_col = find_column(df, ["section"])
    spec_col = find_column(df, ["specialization", "specialisation"])
    admission_col = find_column(df, ["admission type", "lateral entry", "entry type"])

    # Uploaded-record uniqueness
    uploaded_reg: Dict[str, List[int]] = {}
    uploaded_emails: Dict[str, List[int]] = {}
    uploaded_phones: Dict[str, List[int]] = {}

    for idx, row in df.iterrows():
        if reg_col:
            reg = normalise_whitespace(row[reg_col])
            if reg:
                uploaded_reg.setdefault(reg.lower(), []).append(idx)
        if email_col:
            email = normalise_whitespace(row[email_col]).lower()
            if email:
                uploaded_emails.setdefault(email, []).append(idx)
        if phone_col:
            phone = clean_phone(row[phone_col])
            if phone:
                uploaded_phones.setdefault(phone, []).append(idx)

    for idx, row in df.iterrows():
        if _is_blank_row(row):
            remarks.append("")
            continue
        msgs = []

        # Registration Id
        reg = ""
        if reg_col:
            reg = normalise_whitespace(row[reg_col])
            if not reg:
                msgs.append("Registration Id is mandatory")
            else:
                reg_l = reg.lower()
                if reg_l in uploaded_reg and len(uploaded_reg[reg_l]) > 1:
                    msgs.append("Duplicate Registration Id within uploaded records")
                elif not fresh_instance and reg_l in existing_reg_ids:
                    msgs.append("Registration Id already exists in system (user_attributes)")
        else:
            msgs.append("Mandatory field 'Registration Id' is missing in masterdata")

        # Name
        if name_col:
            if is_blank(row[name_col]):
                msgs.append("Name is mandatory")
        else:
            msgs.append("Mandatory field 'Name' is missing in masterdata")

        # Gender
        if gender_col:
            gender = normalise_gender(row[gender_col])
            if not gender:
                msgs.append("Gender must be Male, Female, or Other")
        else:
            msgs.append("Mandatory field 'Gender' is missing in masterdata")

        # Department
        dept = normalise_whitespace(row.get(dept_col, "")) if dept_col else ""
        if not dept:
            msgs.append("Department is mandatory")
        elif dept.lower() not in dept_names:
            msgs.append("Department does not exist in system")

        # Programme
        prog = normalise_whitespace(row.get(prog_col, "")) if prog_col else ""
        if not prog:
            msgs.append("Programme is mandatory")
        elif prog.lower() not in prog_names:
            msgs.append("Programme does not exist in system")

        # Batch Year -> Year of Joining
        year_of_joining = None
        if batch_col:
            batch = normalise_whitespace(row[batch_col])
            if not batch:
                msgs.append("Batch Year is mandatory")
            else:
                year_of_joining = to_int(batch)
                if year_of_joining is None:
                    msgs.append("Batch Year must be a number")
        else:
            msgs.append("Mandatory field 'Batch Year' is missing in masterdata")

        # Intake
        intake = ""
        if intake_col:
            intake = normalise_whitespace(row[intake_col])
        # Derive expected intake from programme + batch year
        expected_intake = ""
        if prog and year_of_joining is not None:
            expected_intake = f"{prog}-{year_of_joining}-intake"
        if not intake:
            if expected_intake:
                intake = expected_intake
            elif intake_col:
                msgs.append("Intake is mandatory")
            else:
                msgs.append("Mandatory field 'Intake' is missing in masterdata")

        if intake:
            # Detect Excel date artifacts such as "2018-06-01 00:00:00"
            looks_like_date = bool(re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", intake))
            if looks_like_date:
                msgs.append("Intake appears to be a date; expected text like 'Programme-Batch-intake'")
            if not fresh_instance:
                intake_key = _intake_key(intake)
                if intake_key not in intake_name_keys:
                    msgs.append("Intake does not exist in programme_batch_intake table")
                elif mapped_intake_keys and intake_key not in mapped_intake_keys:
                    msgs.append("Intake is not mapped to a configured term in term_programme_batch table")

        # Quota
        quota = normalise_whitespace(row.get(quota_col, "")) if quota_col else ""
        if not quota:
            msgs.append("Quota is mandatory")
        elif not fresh_instance and quota.lower() not in quota_names:
            msgs.append("Quota does not exist in system")

        # Email
        email = ""
        if email_col:
            email = normalise_whitespace(row[email_col])
            if not email:
                msgs.append("Email is mandatory")
            elif not is_valid_email(email):
                msgs.append("Invalid Email format")
            else:
                email_l = email.lower()
                if not fresh_instance and email_l in auth_emails:
                    msgs.append("Email already exists in system")
                elif email_l in uploaded_emails and len(uploaded_emails[email_l]) > 1:
                    msgs.append("Duplicate Email within uploaded records")
        else:
            msgs.append("Mandatory field 'Email' is missing in masterdata")

        # Phone
        if phone_col:
            phone = clean_phone(row[phone_col])
            raw_phone = normalise_whitespace(row[phone_col])
            if not raw_phone:
                msgs.append("Phone Number is mandatory")
            elif phone is None:
                msgs.append("Phone Number must be 10 digits")
            else:
                if not fresh_instance and phone in auth_phones:
                    msgs.append("Phone Number already exists in system")
                elif phone in uploaded_phones and len(uploaded_phones[phone]) > 1:
                    msgs.append("Duplicate Phone Number within uploaded records")
        else:
            msgs.append("Mandatory field 'Phone Number' is missing in masterdata")

        # Lateral entry check
        if admission_col:
            admission = normalise_whitespace(row[admission_col]).lower()
            if "lateral" in admission and year_of_joining is None:
                msgs.append("Year of Joining cannot be empty for lateral entry students")

        remarks.append("; ".join(msgs) if msgs else OK_TEXT)

    df[REMARKS_COL] = remarks
    return df


# ---------------------------------------------------------------------------
# Output writer
# ---------------------------------------------------------------------------


def write_output(
    input_path: str,
    output_path: str,
    assessed_sheets: Dict[str, pd.DataFrame],
) -> None:
    """Copy the original workbook and overwrite the assessed sheets with their Remarks column."""
    import openpyxl

    wb = openpyxl.load_workbook(input_path)

    for logical_name, df in assessed_sheets.items():
        sheet_name = SHEET_MAP[logical_name]["sheet"]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Determine header row for this sheet (1-based in openpyxl)
        header_row = SHEET_MAP[logical_name]["header_row"] + 1

        # Add Remarks header
        remarks_col_idx = ws.max_column + 1
        ws.cell(row=header_row, column=remarks_col_idx, value=REMARKS_COL)

        # Write remarks aligned to data rows. pandas iterrows is 0-based; Excel data starts at header_row + 1
        for pandas_idx, remark in enumerate(df[REMARKS_COL], start=0):
            excel_row = header_row + 1 + pandas_idx
            ws.cell(row=excel_row, column=remarks_col_idx, value=remark)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Upload-format generators
# ---------------------------------------------------------------------------


def generate_student_file(
    assessed_sheets: Dict[str, pd.DataFrame],
    output_path: str,
    layer_map: Dict[str, str],
) -> int:
    """Generate the student creation upload file from OK rows only."""
    student_df = assessed_sheets.get("student")
    if student_df is None:
        return 0

    ok_mask = student_df[REMARKS_COL] == OK_TEXT
    ok_df = student_df[ok_mask].copy()

    dept_col = find_column(ok_df, ["department"])
    intake_col = find_column(ok_df, ["intake"])

    # Column mapping: output column -> source column finder(s)
    out_cols = {
        "Registration Id": ["registration id"],
        "Name": ["name"],
        "Layer": [],
        "Department": ["department"],
        "Programme": ["programme", "programme/degree"],
        "Quota": ["quota"],
        "Admission Type(REGULAR/ LATERAL_ENTRY)": [],
        "Year of Joining": ["batch year"],
        "Intake": ["intake"],
        "Email Id": ["email"],
        "Phone Number": ["phone number"],
        "Password": [],
        "Gender(Male/Female/Other)": ["gender"],
        "Section Name": ["section"],
        "Batch Year": ["batch year"],
        "Joining Date(YYYY-MM-DD)": [],
        "Expected Year of Passing": [],
        "Expected Date of Passing(YYYY-MM-DD)": [],
        "Send Account Creation Email(YES/NO)": [],
    }

    # Build the output DataFrame
    rows = []
    for _, row in ok_df.iterrows():
        programme = normalise_whitespace(row.get(find_column(ok_df, ["programme", "programme/degree"]), ""))
        batch_year = to_int(row.get(find_column(ok_df, ["batch year"]), ""))

        dept_name = normalise_whitespace(row.get(dept_col, "")) if dept_col else ""
        layer = layer_map.get(dept_name.lower(), "") if dept_name else ""

        record = {
            "Registration Id": normalise_whitespace(row.get(find_column(ok_df, ["registration id"]), "")),
            "Name": normalise_whitespace(row.get(find_column(ok_df, ["name"]), "")),
            "Layer": layer,
            "Department": dept_name,
            "Programme": programme,
            "Quota": normalise_whitespace(row.get(find_column(ok_df, ["quota"]), "")),
            "Admission Type(REGULAR/ LATERAL_ENTRY)": "REGULAR",
            "Year of Joining": batch_year if batch_year is not None else "",
            "Intake": _format_intake(
                normalise_whitespace(row.get(intake_col, "")) if intake_col else "",
                programme,
                batch_year,
            ),
            "Email Id": normalise_whitespace(row.get(find_column(ok_df, ["email"]), "")),
            "Phone Number": clean_phone(row.get(find_column(ok_df, ["phone number"]), "")) or "",
            "Password": "",
            "Gender(Male/Female/Other)": normalise_gender(row.get(find_column(ok_df, ["gender"]), "")) or "",
            "Section Name": "",
            "Batch Year": batch_year if batch_year is not None else "",
            "Joining Date(YYYY-MM-DD)": "",
            "Expected Year of Passing": "",
            "Expected Date of Passing(YYYY-MM-DD)": "",
            "Send Account Creation Email(YES/NO)": "NO",
        }
        rows.append(record)

    out_df = pd.DataFrame(rows, columns=list(out_cols.keys()))
    out_df.to_excel(output_path, index=False, sheet_name="Sample")
    return len(out_df)


def generate_staff_file(
    assessed_sheets: Dict[str, pd.DataFrame],
    output_path: str,
    layer_map: Dict[str, str],
) -> int:
    """Generate the staff bulk creation upload file from OK faculty + admin rows."""
    fac_df = assessed_sheets.get("faculty")
    admin_df = assessed_sheets.get("admin")

    parts = []
    if fac_df is not None:
        ok_fac = fac_df[fac_df[REMARKS_COL] == OK_TEXT].copy()
        if not ok_fac.empty:
            ok_fac["__user_type__"] = "faculty"
            parts.append(ok_fac)
    if admin_df is not None:
        ok_admin = admin_df[admin_df[REMARKS_COL] == OK_TEXT].copy()
        if not ok_admin.empty:
            ok_admin["__user_type__"] = "administrator"
            parts.append(ok_admin)

    if parts:
        combined = pd.concat(parts, ignore_index=True)
    else:
        combined = pd.DataFrame()

    dept_col = find_column(combined, ["department"])

    rows = []
    for _, row in combined.iterrows():
        gender = normalise_gender(row.get(find_column(combined, ["gender"]), ""))
        if gender:
            gender = gender.lower()

        dept_name = normalise_whitespace(row.get(dept_col, "")) if dept_col else ""
        layer = layer_map.get(dept_name.lower(), "") if dept_name else ""

        record = {
            "User Type*": row.get("__user_type__", ""),
            "Employee ID*": normalise_whitespace(row.get(find_column(combined, ["employee id"]), "")),
            "Name*": normalise_whitespace(row.get(find_column(combined, ["name"]), "")),
            "Gender*": gender or "",
            "Email*": normalise_whitespace(row.get(find_column(combined, ["email"]), "")),
            "Phone Number*": clean_phone(row.get(find_column(combined, ["phone number"]), "")) or "",
            "Layer": layer,
            "Department*": dept_name,
            "Designation": normalise_whitespace(row.get(find_column(combined, ["designation"]), "")),
            "Post Name": "",
            "Reservation Name": "",
            "Salutation": "",
            "Accommodation Type": "NON_RESIDENT",
            "MFA Enabled": "disabled",
        }
        rows.append(record)

    out_cols = [
        "User Type*",
        "Employee ID*",
        "Name*",
        "Gender*",
        "Email*",
        "Phone Number*",
        "Layer",
        "Department*",
        "Designation",
        "Post Name",
        "Reservation Name",
        "Salutation",
        "Accommodation Type",
        "MFA Enabled",
    ]
    out_df = pd.DataFrame(rows, columns=out_cols)
    out_df.to_excel(output_path, index=False, sheet_name="Sample")
    return len(out_df)


# ---------------------------------------------------------------------------
# Interactive prompts
# ---------------------------------------------------------------------------


def prompt_tabs() -> List[str]:
    print("\nSelect tabs to validate (enter numbers separated by commas, or 'all'):")
    for i, key in enumerate(SHEET_MAP, start=1):
        print(f"  {i}. {key}  ->  {SHEET_MAP[key]['sheet']}")

    raw = input("\nSelection: ").strip().lower()
    if raw in {"", "all", "a"}:
        return list(SHEET_MAP.keys())

    selected = []
    parts = [p.strip() for p in raw.split(",")]
    keys = list(SHEET_MAP.keys())
    for p in parts:
        try:
            idx = int(p)
            if 1 <= idx <= len(keys):
                selected.append(keys[idx - 1])
        except ValueError:
            if p in SHEET_MAP:
                selected.append(p)
    return list(dict.fromkeys(selected))  # preserve order, dedupe


def resolve_input_path() -> str:
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        return sys.argv[1].strip().strip('"').strip("'")
    print("Tip: you can drag-and-drop the Excel file onto this script.\n")
    raw = input("Enter the path to the masterdata Excel file: ").strip()
    return raw.strip('"').strip("'")


def resolve_schema() -> str:
    if len(sys.argv) >= 3 and sys.argv[2].strip():
        return sys.argv[2].strip()
    print(f"\nKnown schemas (first 20): {', '.join(list_schemas()[:20])}")
    return input("Enter tenant schema (e.g. collpoll_tenant or tenant): ").strip()


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate BSACIST masterdata Excel against tenant DB")
    parser.add_argument("input", nargs="?", help="Input Excel file path")
    parser.add_argument("schema", nargs="?", help="Tenant schema name")
    parser.add_argument("--tabs", help="Comma-separated logical tab names to assess (e.g. department,programme,faculty,admin,student)")
    parser.add_argument("--output", help="Output Excel file path (default: <input>_validated.xlsx)")
    return parser.parse_args(argv)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def validate_workbook(
    input_path: str,
    output_folder: str,
    instance_type: str,
    schema: Optional[str],
    tabs: List[str],
    progress_callback=None,
) -> List[Tuple[str, int, int]]:
    """Core validation entry point.

    Validates the selected tabs and writes three outputs into `output_folder`:
      - masterdata_validated.xlsx  (original workbook + Remarks column)
      - student_creation.xlsx      (OK student rows in upload format)
      - staff_creation.xlsx        (OK faculty + admin rows in upload format)

    `instance_type` is either "fresh" (cross-tab validation, no DB) or
    "existing" (validate against the tenant database; schema required).

    Raises exceptions on failure; returns a summary list on success.
    `progress_callback(message)` is called with status updates when provided.
    """
    fresh_instance = str(instance_type).strip().lower() == "fresh"

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if not fresh_instance and not schema:
        raise ValueError("Tenant schema is required for existing instances.")

    valid_tabs = [t for t in tabs if t in SHEET_MAP]
    if not valid_tabs:
        raise ValueError("No valid tabs selected.")

    def _log(msg: str):
        if progress_callback:
            progress_callback(msg)
        else:
            print(msg)

    # -----------------------------------------------------------------------
    # Build validation context (DB for existing, masterdata tabs for fresh)
    # -----------------------------------------------------------------------
    conn = None
    assessed_sheets: Dict[str, pd.DataFrame] = {}

    if fresh_instance:
        _log("Fresh instance mode: cross-tab validation (no DB lookup).")
        # Load department and programme first so FreshContext can use them.
        for logical in ("department", "programme"):
            if logical not in valid_tabs:
                continue
            meta = SHEET_MAP[logical]
            xl = pd.ExcelFile(input_path)
            if meta["sheet"] not in xl.sheet_names:
                continue
            assessed_sheets[logical] = load_sheet(input_path, meta["sheet"], header_row=meta["header_row"])
        db: DbCache = FreshContext(assessed_sheets)
    else:
        if _mysql_import_error:
            raise RuntimeError(
                "mysql-connector-python is required but not installed. "
                "Run: .venv/Scripts/python -m pip install mysql-connector-python"
            )
        cfg = get_db_config(schema)
        _log(f"Connecting to {cfg['database']} at {cfg['host']} ...")
        conn = mysql.connector.connect(
            host=cfg["host"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            connection_timeout=15,
        )
        db = DbCache(conn)

        # Pre-fetch required tables based on selected tabs
        required_tables = set()
        if "department" in valid_tabs:
            required_tables.add("department")
        if "programme" in valid_tabs:
            required_tables.update({"department", "programme"})
        if "course" in valid_tabs:
            required_tables.update({"department", "course"})
        if "faculty" in valid_tabs or "admin" in valid_tabs:
            required_tables.update({"department", "institution_entity", "institution_layer", "authenticator"})
        if "student" in valid_tabs:
            required_tables.update({
                "department",
                "institution_entity",
                "institution_layer",
                "programme",
                "quota",
                "programme_batch_intake",
                "term_programme_batch",
                "authenticator",
                "user_attributes",
            })

        for table in required_tables:
            try:
                db.fetch(table)
            except mysql.connector.Error as e:
                _log(f"WARNING: Could not read table '{table}': {e}")

    summary: List[Tuple[str, int, int]] = []

    try:
        # Validate each selected tab
        for logical in valid_tabs:
            meta = SHEET_MAP[logical]
            sheet_name = meta["sheet"]
            header_row = meta["header_row"]

            # Check sheet exists
            xl = pd.ExcelFile(input_path)
            if sheet_name not in xl.sheet_names:
                _log(f"WARNING: sheet '{sheet_name}' not found; skipping {logical}")
                continue

            # For fresh mode, department/programme were pre-loaded; still need to validate them.
            if logical in assessed_sheets:
                df = assessed_sheets[logical]
            else:
                df = load_sheet(input_path, sheet_name, header_row=header_row)

            # Keep blank rows so __excel_row__ alignment is preserved; validators skip them
            data_row_count = len(df) - df.apply(_is_blank_row, axis=1).sum()

            _log(f"Validating '{sheet_name}' ({data_row_count} data rows) ...")

            if logical == "department":
                df = validate_department(df, db, fresh_instance)
            elif logical == "programme":
                df = validate_programme(df, db, fresh_instance)
            elif logical == "course":
                df = validate_course(df, db, fresh_instance)
            elif logical in ("faculty", "admin"):
                # Validate both together for cross-sheet duplicate detection
                if "faculty" not in assessed_sheets and "admin" not in assessed_sheets:
                    fac_meta = SHEET_MAP["faculty"]
                    admin_meta = SHEET_MAP["admin"]
                    fac_df = load_sheet(input_path, fac_meta["sheet"], header_row=fac_meta["header_row"])
                    admin_df = load_sheet(input_path, admin_meta["sheet"], header_row=admin_meta["header_row"])
                    fac_df, admin_df = validate_faculty_and_admin(fac_df, admin_df, db, fresh_instance)
                    assessed_sheets["faculty"] = fac_df
                    assessed_sheets["admin"] = admin_df
                df = assessed_sheets[logical]
            elif logical == "student":
                df = validate_student(df, db, fresh_instance)

            assessed_sheets[logical] = df
            non_blank_mask = ~df.apply(_is_blank_row, axis=1)
            total_data = int(non_blank_mask.sum())
            bad = int(((df[REMARKS_COL] != OK_TEXT) & non_blank_mask).sum())
            summary.append((logical, total_data, bad))
            _log(f"  {logical}: {total_data} data rows, {bad} with remarks")

        # Ensure output folder exists
        os.makedirs(output_folder, exist_ok=True)

        # Write masterdata output
        masterdata_path = os.path.join(output_folder, "masterdata_validated.xlsx")
        write_output(input_path, masterdata_path, assessed_sheets)
        _log(f"Saved validated workbook to: {masterdata_path}")

        # Build department -> layer mapping for upload files
        layer_map = db.department_layer_map()

        # Generate upload-format files from OK rows (files are created even if empty)
        generated: List[str] = []

        if "student" in assessed_sheets:
            student_path = os.path.join(output_folder, "student_creation.xlsx")
            count = generate_student_file(assessed_sheets, student_path, layer_map)
            _log(f"Generated student_creation.xlsx ({count} OK rows)")
            generated.append(student_path)

        if "faculty" in assessed_sheets or "admin" in assessed_sheets:
            staff_path = os.path.join(output_folder, "staff_creation.xlsx")
            count = generate_staff_file(assessed_sheets, staff_path, layer_map)
            _log(f"Generated staff_creation.xlsx ({count} OK rows)")
            generated.append(staff_path)

        _log("Summary:")
        for logical, total, bad in summary:
            _log(f"  {logical:12s}: {total:4d} rows checked, {bad:4d} with remarks")

    finally:
        if conn is not None and conn.is_connected():
            conn.close()

    return summary



# ---------------------------------------------------------------------------
# Tkinter UI
# ---------------------------------------------------------------------------

import os
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError as exc:  # pragma: no cover
    print(f"ERROR: tkinterdnd2 is required for drag-and-drop: {exc}")
    print("Run:  .venv/Scripts/python -m pip install tkinterdnd2")
    sys.exit(1)


class MasterdataValidatorUI:
    def __init__(self, root: TkinterDnD.Tk):
        self.root = root
        self.root.title("Masterdata Validator")
        self.root.geometry("800x700")
        self.root.minsize(700, 550)

        self._build_widgets()
        self._set_default_output()

    def _build_widgets(self):
        pad = {"padx": 12, "pady": 8}

        # ---- File selection --------------------------------------------------
        file_frame = ttk.LabelFrame(self.root, text="Masterdata Excel File")
        file_frame.pack(fill="x", **pad)

        self.file_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_var)
        file_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        browse_btn = ttk.Button(file_frame, text="Browse", command=self._browse_file)
        browse_btn.pack(side="right", padx=(4, 0))

        drop_lbl = tk.Label(
            file_frame,
            text="…or drop a file here",
            fg="gray",
            bg="#f0f0f0",
            relief="ridge",
        )
        drop_lbl.pack(side="right", padx=(4, 0))
        drop_lbl.drop_target_register(DND_FILES)
        drop_lbl.dnd_bind("<<Drop>>", self._on_drop)

        # ---- Instance type ---------------------------------------------------
        instance_frame = ttk.LabelFrame(self.root, text="Instance Type")
        instance_frame.pack(fill="x", **pad)

        self.instance_var = tk.StringVar(value="existing")
        ttk.Radiobutton(
            instance_frame,
            text="Existing (validate against database)",
            variable=self.instance_var,
            value="existing",
            command=self._on_instance_change,
        ).pack(side="left", padx=(0, 16))
        ttk.Radiobutton(
            instance_frame,
            text="Fresh (cross-tab validation only)",
            variable=self.instance_var,
            value="fresh",
            command=self._on_instance_change,
        ).pack(side="left")

        # ---- Tenant / schema -------------------------------------------------
        schema_frame = ttk.LabelFrame(self.root, text="Tenant / Schema")
        schema_frame.pack(fill="x", **pad)

        self.schema_var = tk.StringVar()
        self.schema_combo = ttk.Combobox(
            schema_frame,
            textvariable=self.schema_var,
            values=list_schemas(),
            width=40,
        )
        self.schema_combo.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.schema_combo.set("")

        refresh_btn = ttk.Button(schema_frame, text="Refresh list", command=self._refresh_schemas)
        refresh_btn.pack(side="right")

        # ---- Tabs checkboxes -------------------------------------------------
        tabs_frame = ttk.LabelFrame(self.root, text="Tabs to Validate")
        tabs_frame.pack(fill="x", **pad)

        self.tab_vars: dict[str, tk.BooleanVar] = {}
        cb_frame = ttk.Frame(tabs_frame)
        cb_frame.pack(fill="x", pady=4)
        for i, key in enumerate(SHEET_MAP.keys()):
            var = tk.BooleanVar(value=True)
            self.tab_vars[key] = var
            ttk.Checkbutton(
                cb_frame,
                text=f"{key}  ({SHEET_MAP[key]['sheet']})",
                variable=var,
            ).grid(row=i // 3, column=i % 3, sticky="w", padx=12, pady=2)

        btn_frame = ttk.Frame(tabs_frame)
        btn_frame.pack(fill="x", pady=(4, 0))
        ttk.Button(btn_frame, text="Select All", command=self._select_all).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Clear All", command=self._clear_all).pack(side="left")

        # ---- Output folder ---------------------------------------------------
        out_frame = ttk.LabelFrame(self.root, text="Output Folder")
        out_frame.pack(fill="x", **pad)

        self.output_var = tk.StringVar()
        self.output_entry = ttk.Entry(out_frame, textvariable=self.output_var)
        self.output_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        ttk.Button(out_frame, text="Browse", command=self._browse_output).pack(side="right", padx=(4, 0))
        ttk.Button(out_frame, text="Reset", command=self._set_default_output).pack(side="right", padx=(4, 0))

        # ---- Validate button -------------------------------------------------
        self.validate_btn = ttk.Button(
            self.root,
            text="Validate",
            command=self._on_validate,
        )
        self.validate_btn.pack(pady=(8, 4))

        # ---- Log area --------------------------------------------------------
        log_frame = ttk.LabelFrame(self.root, text="Log")
        log_frame.pack(fill="both", expand=True, **pad)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap="word",
            state="disabled",
            height=15,
        )
        self.log_text.pack(fill="both", expand=True, padx=4, pady=4)

    def _log(self, message: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _on_drop(self, event):
        path = event.data
        path = path.strip().strip('"').strip("'").strip("{}")
        if path:
            self.file_var.set(path)
            self._set_default_output()

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select masterdata Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.file_var.set(path)
            self._set_default_output()

    def _browse_output(self):
        path = filedialog.askdirectory(
            title="Select output folder",
            mustexist=False,
        )
        if path:
            self.output_var.set(path)

    def _set_default_output(self):
        desktop = os.path.join(str(Path.home()), "Desktop")
        input_path = self.file_var.get().strip()
        if input_path:
            base = os.path.splitext(os.path.basename(input_path))[0]
            folder = os.path.join(desktop, base + "_upload_ready")
            self.output_var.set(folder)
        else:
            self.output_var.set(os.path.join(desktop, "masterdata_upload_ready"))

    def _on_instance_change(self):
        if self.instance_var.get() == "fresh":
            self.schema_combo.configure(state="disabled")
        else:
            self.schema_combo.configure(state="normal")

    def _refresh_schemas(self):
        self.schema_combo["values"] = list_schemas()
        self._log("Schema list refreshed.")

    def _select_all(self):
        for var in self.tab_vars.values():
            var.set(True)

    def _clear_all(self):
        for var in self.tab_vars.values():
            var.set(False)

    def _on_validate(self):
        input_path = self.file_var.get().strip()
        instance_type = self.instance_var.get().strip()
        schema = self.schema_var.get().strip()
        output_folder = self.output_var.get().strip()

        if not input_path:
            messagebox.showerror("Missing file", "Please select or drop the masterdata Excel file.")
            return
        if not os.path.exists(input_path):
            messagebox.showerror("File not found", f"File not found:\n{input_path}")
            return
        if instance_type == "existing" and not schema:
            messagebox.showerror("Missing schema", "Please enter or select a tenant schema for existing instances.")
            return
        if not output_folder:
            messagebox.showerror("Missing output", "Please specify an output folder path.")
            return

        tabs = [key for key, var in self.tab_vars.items() if var.get()]
        if not tabs:
            messagebox.showerror("No tabs", "Please select at least one tab to validate.")
            return

        self._clear_log()
        self.validate_btn.configure(state="disabled")
        self._log("Starting validation...")

        def run():
            try:
                summary = validate_workbook(
                    input_path,
                    output_folder,
                    instance_type,
                    schema if instance_type == "existing" else None,
                    tabs,
                    progress_callback=lambda msg: self.root.after(0, self._log, msg),
                )
                self.root.after(0, self._on_success, output_folder, summary)
            except Exception as exc:
                self.root.after(0, self._on_error, exc)

        thread = threading.Thread(target=run, daemon=True)
        thread.start()

    def _on_success(self, output_folder: str, summary: list):
        self.validate_btn.configure(state="normal")
        lines = ["Validation complete.", f"Output folder: {output_folder}", ""]
        lines.append("Summary:")
        for logical, total, bad in summary:
            lines.append(f"  {logical:12s}: {total:4d} rows checked, {bad:4d} with remarks")
        self._log("\n".join(lines))
        messagebox.showinfo("Validation Complete", f"Output files saved to:\n{output_folder}")

    def _on_error(self, exc: Exception):
        self.validate_btn.configure(state="normal")
        msg = f"ERROR: {exc}"
        self._log(msg)
        messagebox.showerror("Validation Failed", str(exc))


def main():
    root = TkinterDnD.Tk()
    app = MasterdataValidatorUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
