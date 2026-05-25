from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
import io

app = Flask(__name__)

# Store form data across steps
form_data = {
    'servicedesk': {},
    'service': [],
    'form': [],
    'fields': [],
    'option_for_dropdown': [],
    'attachment': [],
    'workcentre': [],
    'workcentre_action': [],
    'workcentre_action_form': [],
    'workcentre_action_form_fields': [],
    'workcentre_action_form_attachment': []
}

@app.route('/')
def index():
    return redirect(url_for('servicedesk'))

@app.route('/servicedesk', methods=['GET', 'POST'])
def servicedesk():
    if request.method == 'POST':
        form_data['servicedesk'] = {
            'title': request.form.get('title'),
            'icon': request.form.get('icon', 'default'),
            'parentServiceDeskName': request.form.get('parentServiceDeskName', ''),
            'parentServiceDeskId': request.form.get('parentServiceDeskId', '')
        }
        return redirect(url_for('service'))
    return render_template('servicedesk.html')

@app.route('/service', methods=['GET', 'POST'])
def service():
    if request.method == 'POST':
        if 'add_service' in request.form:
            form_data['service'].append({
                'id': request.form.get('id', ''),
                'title': request.form.get('title'),
                'description': request.form.get('description', ''),
                'icon': request.form.get('icon', 'default'),
                'serviceDeskName': request.form.get('serviceDeskName', ''),
                'serviceDeskId': request.form.get('serviceDeskId', ''),
                'paid': request.form.get('paid', ''),
                'amount': request.form.get('amount', ''),
                'currency': request.form.get('currency', ''),
                'costItemId': request.form.get('costItemId', ''),
                'escalationEnabled': request.form.get('escalationEnabled', ''),
                'commentsDisabled': request.form.get('commentsDisabled', ''),
                'deleted': request.form.get('deleted', ''),
                'anonymous': request.form.get('anonymous', ''),
                'sendEmail': request.form.get('sendEmail', ''),
                'duplicateRequest': request.form.get('duplicateRequest', ''),
                'isFeedbackEnabled': request.form.get('isFeedbackEnabled', ''),
                'workingHoursEnabled': request.form.get('workingHoursEnabled', ''),
                'isMiddlewareService': request.form.get('isMiddlewareService', 'false')
            })
            return redirect(url_for('service'))
        elif 'next' in request.form:
            return redirect(url_for('form'))
    return render_template('service.html', services=form_data['service'])

@app.route('/form', methods=['GET', 'POST'])
def form():
    if request.method == 'POST':
        if 'add_form' in request.form:
            form_data['form'].append({
                'title': request.form.get('title'),
                'description': request.form.get('description', ''),
                'autoPopulateEmail': request.form.get('autoPopulateEmail', ''),
                'autoPopulatePhone': request.form.get('autoPopulatePhone', ''),
                'autoPopulateHostelDetails': request.form.get('autoPopulateHostelDetails', ''),
                'disclaimerRequired': request.form.get('disclaimerRequired', 'false'),
                'disclaimer': request.form.get('disclaimer', ''),
                'serviceName': request.form.get('serviceName', ''),
                'serviceId': request.form.get('serviceId', ''),
                'showEmailToResolver': request.form.get('showEmailToResolver', ''),
                'showPhoneToResolver': request.form.get('showPhoneToResolver', ''),
                'showHostelDetailsToResolver': request.form.get('showHostelDetailsToResolver', '')
            })
            return redirect(url_for('form'))
        elif 'next' in request.form:
            return redirect(url_for('fields'))
    return render_template('form.html', forms=form_data['form'], services=form_data['service'])

@app.route('/fields', methods=['GET', 'POST'])
def fields():
    if request.method == 'POST':
        if 'add_field' in request.form:
            form_data['fields'].append({
                'id': request.form.get('id', ''),
                'name': request.form.get('name'),
                'element': request.form.get('element', 'textField'),
                'minimumLength': request.form.get('minimumLength', ''),
                'maximumLength': request.form.get('maximumLength', ''),
                'minimumNumber': request.form.get('minimumNumber', ''),
                'maximumNumber': request.form.get('maximumNumber', ''),
                'mandatory': request.form.get('mandatory', '0'),
                'formName': request.form.get('formName', ''),
                'formId': request.form.get('formId', ''),
                'showToResolver': request.form.get('showToResolver', '0'),
                'fieldTypeParent': request.form.get('fieldTypeParent', 'null'),
                'userFieldIdentifier': request.form.get('userFieldIdentifier', 'null')
            })
            return redirect(url_for('fields'))
        elif 'next' in request.form:
            return redirect(url_for('option_for_dropdown'))
    return render_template('fields.html', fields=form_data['fields'], forms=form_data['form'])

@app.route('/option_for_dropdown', methods=['GET', 'POST'])
def option_for_dropdown():
    if request.method == 'POST':
        if 'add_option' in request.form:
            form_data['option_for_dropdown'].append({
                'fieldId': request.form.get('fieldId', ''),
                'label': request.form.get('label')
            })
            return redirect(url_for('option_for_dropdown'))
        elif 'next' in request.form:
            return redirect(url_for('attachment'))
    return render_template('option_for_dropdown.html', options=form_data['option_for_dropdown'], fields=form_data['fields'])

@app.route('/attachment', methods=['GET', 'POST'])
def attachment():
    if request.method == 'POST':
        if 'add_attachment' in request.form:
            form_data['attachment'].append({
                'attachmentLabel': request.form.get('attachmentLabel'),
                'formName': request.form.get('formName'),
                'formId': request.form.get('formId', ''),
                'mandatory': request.form.get('mandatory', '0')
            })
            return redirect(url_for('attachment'))
        elif 'next' in request.form:
            return redirect(url_for('workcentre'))
    return render_template('attachment.html', attachments=form_data['attachment'], forms=form_data['form'])

@app.route('/workcentre', methods=['GET', 'POST'])
def workcentre():
    if request.method == 'POST':
        if 'add_workcentre' in request.form:
            form_data['workcentre'].append({
                'title': request.form.get('title'),
                'assignmentType': request.form.get('assignmentType', 'defaultAssign'),
                'reassignmentType': request.form.get('reassignmentType', 'manualAssign'),
                'parentWorkcentreName': request.form.get('parentWorkcentreName', ''),
                'parentWorkcentreId': request.form.get('parentWorkcentreId', ''),
                'closureWorkcentre': request.form.get('closureWorkcentre', '0'),
                'serviceName': request.form.get('serviceName', ''),
                'serviceId': request.form.get('serviceId', ''),
                'chcAssignmentEntity': request.form.get('chcAssignmentEntity', '')
            })
            return redirect(url_for('workcentre'))
        elif 'next' in request.form:
            return redirect(url_for('workcentre_action'))
    return render_template('workcentre.html', workcentres=form_data['workcentre'], services=form_data['service'])

@app.route('/workcentre_action', methods=['GET', 'POST'])
def workcentre_action():
    if request.method == 'POST':
        if 'add_action' in request.form:
            form_data['workcentre_action'].append({
                'id': request.form.get('id', ''),
                'name': request.form.get('name'),
                'requireNote': request.form.get('requireNote', ''),
                'actionTakenName': request.form.get('actionTakenName', ''),
                'defaultAction': request.form.get('defaultAction', '0'),
                'type': request.form.get('type', 'positive'),
                'workcentreName': request.form.get('workcentreName'),
                'workcentreId': request.form.get('workcentreId', ''),
                'formAvailable': request.form.get('formAvailable', '0'),
                'terminationAction': request.form.get('terminationAction', '0')
            })
            return redirect(url_for('workcentre_action'))
        elif 'next' in request.form:
            return redirect(url_for('workcentre_action_form'))
    return render_template('workcentre_action.html', actions=form_data['workcentre_action'], workcentres=form_data['workcentre'])

@app.route('/workcentre_action_form', methods=['GET', 'POST'])
def workcentre_action_form():
    if request.method == 'POST':
        if 'add_action_form' in request.form:
            form_data['workcentre_action_form'].append({
                'title': request.form.get('title'),
                'description': request.form.get('description', ''),
                'autoPopulateEmail': request.form.get('autoPopulateEmail', ''),
                'autoPopulatePhone': request.form.get('autoPopulatePhone', ''),
                'actionId': request.form.get('actionId', ''),
                'showRequester': request.form.get('showRequester', '0')
            })
            return redirect(url_for('workcentre_action_form'))
        elif 'next' in request.form:
            return redirect(url_for('workcentre_action_form_fields'))
    return render_template('workcentre_action_form.html', action_forms=form_data['workcentre_action_form'], actions=form_data['workcentre_action'])

@app.route('/workcentre_action_form_fields', methods=['GET', 'POST'])
def workcentre_action_form_fields():
    if request.method == 'POST':
        if 'add_field' in request.form:
            form_data['workcentre_action_form_fields'].append({
                'name': request.form.get('name'),
                'element': request.form.get('element', 'textField'),
                'minimum_length': request.form.get('minimum_length', ''),
                'maximum_length': request.form.get('maximum_length', ''),
                'minimum_number': request.form.get('minimum_number', ''),
                'maximum_number': request.form.get('maximum_number', ''),
                'mandatory': request.form.get('mandatory', '0'),
                'formName': request.form.get('formName'),
                'formId': request.form.get('formId', '')
            })
            return redirect(url_for('workcentre_action_form_fields'))
        elif 'next' in request.form:
            return redirect(url_for('workcentre_action_form_attachment'))
    return render_template('workcentre_action_form_fields.html', fields=form_data['workcentre_action_form_fields'], action_forms=form_data['workcentre_action_form'])

@app.route('/workcentre_action_form_attachment', methods=['GET', 'POST'])
def workcentre_action_form_attachment():
    if request.method == 'POST':
        if 'add_attachment' in request.form:
            form_data['workcentre_action_form_attachment'].append({
                'attachmentLabel': request.form.get('attachmentLabel'),
                'formName': request.form.get('formName'),
                'formId': request.form.get('formId', ''),
                'mandatory': request.form.get('mandatory', '0')
            })
            return redirect(url_for('workcentre_action_form_attachment'))
        elif 'generate' in request.form:
            return generate_excel()
    return render_template('workcentre_action_form_attachment.html', attachments=form_data['workcentre_action_form_attachment'], action_forms=form_data['workcentre_action_form'])

def generate_excel():
    # Create a new Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Add servicedesk sheet
    ws = wb.create_sheet('servicedesk')
    ws.append(['title', 'icon', 'parentServiceDeskName', 'parentServiceDeskId'])
    if form_data['servicedesk']:
        ws.append([
            form_data['servicedesk']['title'],
            form_data['servicedesk']['icon'],
            form_data['servicedesk']['parentServiceDeskName'],
            form_data['servicedesk']['parentServiceDeskId']
        ])
    
    # Add service sheet
    ws = wb.create_sheet('service')
    ws.append(['id', 'title', 'description', 'icon', 'serviceDeskName', 'serviceDeskId', 'paid', 'amount', 
               'currency', 'costItemId', 'escalationEnabled', 'commentsDisabled', 'deleted', 'anonymous', 
               'sendEmail', 'duplicateRequest', 'isFeedbackEnabled', 'workingHoursEnabled', 'isMiddlewareService'])
    for service in form_data['service']:
        ws.append([
            service['id'], service['title'], service['description'], service['icon'], 
            service['serviceDeskName'], service['serviceDeskId'], service['paid'], 
            service['amount'], service['currency'], service['costItemId'], 
            service['escalationEnabled'], service['commentsDisabled'], service['deleted'], 
            service['anonymous'], service['sendEmail'], service['duplicateRequest'], 
            service['isFeedbackEnabled'], service['workingHoursEnabled'], service['isMiddlewareService']
        ])
    
    # Add form sheet
    ws = wb.create_sheet('form')
    ws.append(['title', 'description', 'autoPopulateEmail', 'autoPopulatePhone', 'autoPopulateHostelDetails', 
               'disclaimerRequired', 'disclaimer', 'serviceName', 'serviceId', 'showEmailToResolver', 
               'showPhoneToResolver', 'showHostelDetailsToResolver'])
    for form in form_data['form']:
        ws.append([
            form['title'], form['description'], form['autoPopulateEmail'], 
            form['autoPopulatePhone'], form['autoPopulateHostelDetails'], 
            form['disclaimerRequired'], form['disclaimer'], form['serviceName'], 
            form['serviceId'], form['showEmailToResolver'], form['showPhoneToResolver'], 
            form['showHostelDetailsToResolver']
        ])
    
    # Add fields sheet
    ws = wb.create_sheet('fields')
    ws.append(['id', 'name', 'element', 'minimumLength', 'maximumLength', 'minimumNumber', 
               'maximumNumber', 'mandatory', 'formName', 'formId', 'showToResolver', 
               'fieldTypeParent', 'userFieldIdentifier'])
    for field in form_data['fields']:
        ws.append([
            field['id'], field['name'], field['element'], field['minimumLength'], 
            field['maximumLength'], field['minimumNumber'], field['maximumNumber'], 
            field['mandatory'], field['formName'], field['formId'], field['showToResolver'], 
            field['fieldTypeParent'], field['userFieldIdentifier']
        ])
    
    # Add option for dropdown sheet
    ws = wb.create_sheet('option for dropdown')
    ws.append(['fieldId', 'label'])
    for option in form_data['option_for_dropdown']:
        ws.append([option['fieldId'], option['label']])
    
    # Add attachment sheet
    ws = wb.create_sheet('attachment')
    ws.append(['attachmentLabel', 'formName', 'formId', 'mandatory'])
    for attachment in form_data['attachment']:
        ws.append([
            attachment['attachmentLabel'], attachment['formName'], 
            attachment['formId'], attachment['mandatory']
        ])
    
    # Add workcentre sheet
    ws = wb.create_sheet('workcentre')
    ws.append(['title', 'assignmentType', 'reassignmentType', 'parentWorkcentreName', 
               'parentWorkcentreId', 'closureWorkcentre', 'serviceName', 'serviceId', 
               'chcAssignmentEntity'])
    for workcentre in form_data['workcentre']:
        ws.append([
            workcentre['title'], workcentre['assignmentType'], workcentre['reassignmentType'], 
            workcentre['parentWorkcentreName'], workcentre['parentWorkcentreId'], 
            workcentre['closureWorkcentre'], workcentre['serviceName'], workcentre['serviceId'], 
            workcentre['chcAssignmentEntity']
        ])
    
    # Add workcentre Action sheet
    ws = wb.create_sheet('workcentre Action')
    ws.append(['id', 'name', 'requireNote', 'actionTakenName', 'defaultAction', 'type', 
               'workcentreName', 'workcentreId', 'formAvailable', 'terminationAction'])
    for action in form_data['workcentre_action']:
        ws.append([
            action['id'], action['name'], action['requireNote'], action['actionTakenName'], 
            action['defaultAction'], action['type'], action['workcentreName'], 
            action['workcentreId'], action['formAvailable'], action['terminationAction']
        ])
    
    # Add work centre action form sheet
    ws = wb.create_sheet('work centre action form')
    ws.append(['title', 'description', 'autoPopulateEmail', 'autoPopulatePhone', 'actionId', 'showRequester'])
    for action_form in form_data['workcentre_action_form']:
        ws.append([
            action_form['title'], action_form['description'], action_form['autoPopulateEmail'], 
            action_form['autoPopulatePhone'], action_form['actionId'], action_form['showRequester']
        ])
    
    # Add work centre action form fields sheet
    ws = wb.create_sheet('work centre action form fields')
    ws.append(['name', 'element', 'minimum length', 'maximum length', 'minimum number', 
               'maximum number', 'mandatory', 'formName', 'formId'])
    for field in form_data['workcentre_action_form_fields']:
        ws.append([
            field['name'], field['element'], field['minimum_length'], field['maximum_length'], 
            field['minimum_number'], field['maximum_number'], field['mandatory'], 
            field['formName'], field['formId']
        ])
    
    # Add work centre action form attachm sheet
    ws = wb.create_sheet('work centre action form attachm')
    ws.append(['attachmentLabel', 'formName', 'formId', 'mandatory'])
    for attachment in form_data['workcentre_action_form_attachment']:
        ws.append([
            attachment['attachmentLabel'], attachment['formName'], 
            attachment['formId'], attachment['mandatory']
        ])
    
    # Save to a bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name='CHC_Configuration.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True)