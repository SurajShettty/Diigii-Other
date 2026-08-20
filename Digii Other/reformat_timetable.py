"""
Reformat & validate a timetable-upload Excel file.

What it does
------------
1. Normalises:
     - Start Date           -> text "dd/mm/yyyy"   (stops Excel re-flipping to mm/dd/yyyy)
     - Start Time / End Time -> text "HH:MM" (24 hr)
2. Validates each row:
     - Class Id*  : not empty AND numeric
     - Start Date : not empty (and parseable)
     - Start Time : not empty (and parseable)
     - End Time   : not empty (and parseable)
3. Any problems for a row are written to a new "Error" column at the far right.
4. The date/time columns in the output are set to TEXT number-format so Excel
   leaves them exactly as written.

Usage
-----
    python reformat_timetable.py input.xlsx
    python reformat_timetable.py input.xlsx output.xlsx

Requires: openpyxl   ->   pip install openpyxl
"""

import sys
import os
import re
from datetime import datetime, time, date

import openpyxl


# ---- file paths -------------------------------------------------------------
# Edit these two placeholders, OR pass paths on the command line which override
# them:  python reformat_timetable.py <input.xlsx> [output.xlsx]
INPUT_PATH  = r"C:\Users\suraj\Downloads\B.com and Mcom (1).xlsx"
# Leave OUTPUT_PATH as None to auto-name it after the input, e.g.
#   "B.com and Mcom (1).xlsx"  ->  "B.com and Mcom (1)_cleaned.xlsx"
# (saved next to the input). Set an explicit path here to override.
OUTPUT_PATH = None


# ---- date-recovery behaviour ------------------------------------------------
# Your input is meant to be dd/mm/yyyy. Under a US (mm/dd) locale Excel can
# silently flip ambiguous dates when it converts them to real datetime cells.
# When SWAP_FLIPPED_DATES is True, any cell Excel stored as a real datetime
# whose day component is <= 12 is treated as a flip and the month/day are
# swapped back to restore dd/mm intent. Each swap is flagged in the Error column
# (prefixed "NOTE:") so you can sanity-check it. Set to False to disable.
SWAP_FLIPPED_DATES = True


# ---- column headers as they appear in row 1 of the template -----------------
COL_CLASS_ID   = "Class Id"
COL_START_DATE = "Start Date"
COL_START_TIME = "Start Time"
COL_END_TIME   = "End Time"


def find_col(headers, wanted):
    """Return the 1-based column index whose header *starts with* `wanted`.

    Headers in the template carry suffixes like '*' or '(dd/mm/yyyy)', so we
    match on a normalised prefix rather than exact text.
    """
    w = wanted.strip().lower()
    for idx, h in enumerate(headers, start=1):
        if h is not None and str(h).strip().lower().startswith(w):
            return idx
    return None


def parse_date(value):
    """Return (date_or_None, note).

    - Real datetime/date cells were converted by Excel and may have been flipped
      from dd/mm to mm/dd. If SWAP_FLIPPED_DATES is on and the day component is
      <= 12 (so the swap is even possible), month and day are swapped back to
      restore dd/mm intent; `note` records what happened.
    - String cells were never converted by Excel, so they are parsed day-first
      (dd/mm/yyyy) as the template expects.
    `note` is None when nothing noteworthy happened.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None, None

    if isinstance(value, (datetime, date)):
        d = value.date() if isinstance(value, datetime) else value
        if SWAP_FLIPPED_DATES and d.day <= 12:
            stored = d.strftime("%d/%m/%Y")
            # swap: new month = old day, new day = old month  -> restores dd/mm
            swapped = date(d.year, d.day, d.month)
            note = (f"date day/month swapped to {swapped.strftime('%d/%m/%Y')} "
                    f"(Excel stored {stored}; ambiguous - verify)")
            return swapped, note
        return d, None

    s = str(value).strip()
    # split on / . or -  e.g. 15/06/2026 , 15-06-2026 , 15.06.2026
    parts = re.split(r"[/\-.]", s)
    if len(parts) == 3:
        try:
            d, m, y = (int(p) for p in parts)
            if y < 100:               # 2-digit year -> 20xx
                y += 2000
            return date(y, m, d), None    # day-first interpretation
        except ValueError:
            return None, None
    # fall back to a couple of explicit day-first formats
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date(), None
        except ValueError:
            continue
    return None, None


def parse_time(value):
    """Return a datetime.time from a cell value, or None if unusable."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value

    s = str(value).strip()
    # Excel sometimes stores a time as a fraction of a day (0.0 - 1.0)
    try:
        f = float(s)
        if 0 <= f < 2:
            total_minutes = round(f * 24 * 60)
            return time(hour=(total_minutes // 60) % 24, minute=total_minutes % 60)
    except ValueError:
        pass

    s_clean = s.upper().replace(".", ":").strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p", "%I%p"):
        try:
            return datetime.strptime(s_clean, fmt).time()
        except ValueError:
            continue
    return None


def parse_class_id(value):
    """Return (int_value, ok). ok is False when empty or non-numeric."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None, False
    if isinstance(value, bool):                 # bools are ints in python - reject
        return None, False
    if isinstance(value, int):
        return value, True
    if isinstance(value, float):
        if value.is_integer():
            return int(value), True
        return None, False
    s = str(value).strip()
    if re.fullmatch(r"\d+", s):
        return int(s), True
    # allow "877.0"
    try:
        f = float(s)
        if f.is_integer():
            return int(f), True
    except ValueError:
        pass
    return None, False


def resolve_input_path():
    """Work out which file to process.

    Priority:
      1. A path given on the command line / dragged onto the .exe (argv[1]).
      2. Otherwise, always ask the user to type or drag-in the path.
    """
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        return sys.argv[1].strip().strip('"').strip("'")
    # always ask (great for a double-clicked .exe)
    print("Tip: you can also drag your Excel file onto this program's icon.\n")
    raw = input("Enter (or drag-and-drop) the path to the Excel file: ").strip()
    # Windows wraps drag-dropped paths in quotes - strip them.
    return raw.strip('"').strip("'")


def process_workbook(in_path, out_path):
    """Reformat & validate the workbook at `in_path`, saving to `out_path`.

    Returns (total_rows, bad_rows). Raises ValueError if the expected columns
    cannot be found in row 1 (message lists the missing headers).
    """
    wb = openpyxl.load_workbook(in_path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    ci_class = find_col(headers, COL_CLASS_ID)
    ci_sdate = find_col(headers, COL_START_DATE)
    ci_stime = find_col(headers, COL_START_TIME)
    ci_etime = find_col(headers, COL_END_TIME)

    missing = [name for name, idx in [
        (COL_CLASS_ID, ci_class), (COL_START_DATE, ci_sdate),
        (COL_START_TIME, ci_stime), (COL_END_TIME, ci_etime)] if idx is None]
    if missing:
        raise ValueError(
            "Could not find these columns in row 1: " + ", ".join(missing)
            + f"\nFound headers: {headers}")

    # add the Error column at the far right
    err_col = ws.max_column + 1
    ws.cell(row=1, column=err_col, value="Error")

    TEXT_FMT = "@"   # forces Excel to treat the cell as plain text
    bad_rows = 0
    total_rows = 0

    for r in range(2, ws.max_row + 1):
        # skip fully blank rows
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ws.max_column + 1)):
            continue
        total_rows += 1
        errors = []
        notes = []

        # --- Class Id -------------------------------------------------------
        cid, ok = parse_class_id(ws.cell(row=r, column=ci_class).value)
        if not ok:
            errors.append("Class Id empty or not numeric")
        else:
            cell = ws.cell(row=r, column=ci_class, value=cid)
            cell.number_format = "0"

        # --- Start Date -----------------------------------------------------
        d, date_note = parse_date(ws.cell(row=r, column=ci_sdate).value)
        if d is None:
            errors.append("Start Date empty or invalid")
        else:
            cell = ws.cell(row=r, column=ci_sdate, value=d.strftime("%d/%m/%Y"))
            cell.number_format = TEXT_FMT
            if date_note:
                notes.append(date_note)

        # --- Start Time -----------------------------------------------------
        st = parse_time(ws.cell(row=r, column=ci_stime).value)
        if st is None:
            errors.append("Start Time empty or invalid")
        else:
            cell = ws.cell(row=r, column=ci_stime, value=st.strftime("%H:%M"))
            cell.number_format = TEXT_FMT

        # --- End Time -------------------------------------------------------
        et = parse_time(ws.cell(row=r, column=ci_etime).value)
        if et is None:
            errors.append("End Time empty or invalid")
        else:
            cell = ws.cell(row=r, column=ci_etime, value=et.strftime("%H:%M"))
            cell.number_format = TEXT_FMT

        # errors (real problems) and notes (informational, e.g. date swaps)
        # both go in the Error column; notes are prefixed "NOTE:".
        messages = list(errors) + [f"NOTE: {n}" for n in notes]
        if errors:
            bad_rows += 1
        if messages:
            ws.cell(row=r, column=err_col, value="; ".join(messages))

    wb.save(out_path)
    return total_rows, bad_rows


def main():
    in_path  = resolve_input_path()
    out_path = sys.argv[2] if len(sys.argv) >= 3 else OUTPUT_PATH

    if not in_path or not os.path.exists(in_path):
        print(f"ERROR: input file not found: {in_path}")
        return

    # no explicit output -> same name as input with a "_cleaned" suffix,
    # saved next to the input file.
    if not out_path:
        base, ext = os.path.splitext(in_path)
        out_path = base + "_cleaned" + (ext or ".xlsx")

    try:
        total_rows, bad_rows = process_workbook(in_path, out_path)
    except ValueError as exc:
        print(f"ERROR: {exc}")
        sys.exit(1)

    print(f"Processed {total_rows} data row(s).")
    print(f"  Rows with errors : {bad_rows}")
    print(f"  Clean rows       : {total_rows - bad_rows}")
    print(f"Saved -> {out_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        # show the problem instead of flashing the window shut
        print(f"\nSomething went wrong: {exc}")
        if "Permission" in type(exc).__name__ or "Permission" in str(exc):
            print("Tip: close the output file in Excel if it's open, then try again.")
    # keep the console window open when launched by double-click / drag-drop
    try:
        input("\nDone. Press Enter to close this window...")
    except EOFError:
        pass
