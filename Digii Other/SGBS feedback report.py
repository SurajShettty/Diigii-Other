import pandas as pd
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path
import re

# ==========================================================
# 1. INPUT SESSION ID
# ==========================================================

session_id = int(input("Enter Session ID: "))

# ==========================================================
# 2. DATABASE CONNECTION
# ==========================================================

conn = mysql.connector.connect(
    host="collpolldb19-read.c5sc77nejhmr.ap-south-1.rds.amazonaws.com",
    user="suraj_shetty",
    password="LW3J0MU3mZ",
    database="collpoll_sgbs"
)

# ==========================================================
# 3. QUERY 1 (RAW DATA - NO GROUP BY)
# ==========================================================

query1 = f"""
SELECT 
fs.session_name as 'Session Name',
cl.batch as 'Class Name',
cl.id as class_id,
concat(ua.f_name, ' ', ua.l_name) as 'Faculty Name',
ua.registration_id as 'Faculty Reg ID',
c.course_code as 'Course Code',
c.course_name as 'Course Name',
cl.type as 'Course Type',
term.name as 'Term',
ft.template_name as 'Template Name',
fts.section_name as 'Section Name',
ftsq.question_text as 'Question',
ftsq.question_type as 'Question Type',
CONCAT('S', LPAD(MOD(CRC32(CONCAT(sp.ukid, pr.programme_id, d.department_id)), 1000000), 6, '0')) as 'Masked Student ID',
coalesce(ftqo.option_text,fsr.response_text) as 'Response/Option',
ftqo.option_score as 'Option Score'
FROM feedback_student_response fsr
LEFT JOIN feedback_student_template_status fsts 
    ON fsts.id = fsr.session_student_template_status_id
LEFT JOIN feedback_template_question_option ftqo 
    ON ftqo.id = fsr.option_id
LEFT JOIN feedback_template_section_question ftsq 
    ON ftsq.id = fsr.question_id
LEFT JOIN feedback_template_section fts 
    ON fts.id = ftsq.template_section_id
LEFT JOIN feedback_session_student fss 
    ON fsts.student_session_id = fss.id
LEFT JOIN feedback_session fs 
    ON fs.id = fss.session_id
LEFT JOIN feedback_course_faculty_template fcft 
    ON fcft.id = fsts.feedback_course_faculty_template_id
LEFT JOIN class cl 
    ON cl.id = fcft.class_id
LEFT JOIN faculty_profile fp 
    ON fp.ukid = fcft.faculty_ukid
LEFT JOIN user_attributes ua 
    ON fp.ukid = ua.ukid
LEFT JOIN feedback_template ft 
    ON ft.id = fcft.template_id
LEFT JOIN course c 
    ON c.course_id = cl.course_id
LEFT JOIN student_profile sp 
    ON sp.ukid = fss.student_ukid
LEFT JOIN programme pr 
    ON pr.programme_id = sp.programme_id
LEFT JOIN department d 
    ON d.department_id = sp.department_id
LEFT JOIN term 
    ON term.id = cl.term_id
WHERE fss.session_id = {session_id}
AND fsts.submitted = 1
"""

raw_df = pd.read_sql(query1, conn)

if raw_df.empty:
    print("No data found.")
    exit()

# ==========================================================
# 4. QUERY 2 (CLASS STRENGTH)
# ==========================================================

query2 = """
SELECT class_id, COUNT(*) as student_count
FROM class_student
GROUP BY class_id
"""
class_strength = pd.read_sql(query2, conn)

# ==========================================================
# 5. SPLIT OBJECTIVE / SUBJECTIVE
# ==========================================================

objective_df = raw_df[raw_df["Question Type"] != "SUBJECTIVE"].copy()
subjective_df = raw_df[raw_df["Question Type"] == "SUBJECTIVE"].copy()

# ==========================================================
# 6. OBJECTIVE CALCULATIONS
# ==========================================================

max_score = objective_df.groupby("Question")["Option Score"].max().reset_index()
max_score.columns = ["Question", "Max Score"]

agg_df = (
    objective_df.groupby([
        "Session Name","Class Name","class_id","Faculty Name",
        "Faculty Reg ID","Course Code","Course Name","Course Type",
        "Term","Template Name","Section Name","Question"
    ])
    .agg(
        Mean=("Option Score","mean"),
        Total_Responses=("Masked Student ID","nunique")
    )
    .reset_index()
)

agg_df = agg_df.merge(max_score,on="Question",how="left")
agg_df["Percentage"] = (agg_df["Mean"] / agg_df["Max Score"]) * 100
agg_df = agg_df.merge(class_strength,on="class_id",how="left")

# SECTION
section_df = (
    agg_df.groupby([
        "Session Name","Class Name","class_id","Faculty Name",
        "Faculty Reg ID","Course Code","Course Name","Course Type",
        "Term","Template Name","Section Name"
    ])
    .agg(
        Section_Mean=("Mean","mean"),
        Section_Percentage=("Percentage","mean")
    )
    .reset_index()
)

# OVERALL
overall_df = (
    agg_df.groupby([
        "Session Name","Class Name","class_id","Faculty Name",
        "Faculty Reg ID","Course Code","Course Name","Course Type",
        "Term","Template Name"
    ])
    .agg(
        Overall_Mean=("Mean","mean"),
        Overall_Percentage=("Percentage","mean")
    )
    .reset_index()
)

# ==========================================================
# 7. CREATE EXCEL
# ==========================================================

wb = Workbook()
ws = wb.active
ws.title = "Objective Feedback"

bold = Font(bold=True)
center = Alignment(horizontal="center",vertical="center",wrap_text=True)
thin = Border(
    left=Side(style='thin'),right=Side(style='thin'),
    top=Side(style='thin'),bottom=Side(style='thin')
)

static_cols = [
    "Session Name","Class Name","Faculty Name","Faculty Reg ID",
    "Course Code","Course Name","Course Type","Term",
    "Template Name","student_count","Total_Responses"
]

# STATIC HEADER
for col_index, name in enumerate(static_cols,1):
    ws.merge_cells(start_row=1,start_column=col_index,
                   end_row=3,end_column=col_index)
    ws.cell(row=1,column=col_index).value=name
    ws.cell(row=1,column=col_index).font=bold
    ws.cell(row=1,column=col_index).alignment=center

col_pointer = len(static_cols) + 1
sections = agg_df["Section Name"].unique()

# SECTION HEADERS
for section in sections:
    sec_questions = agg_df[agg_df["Section Name"]==section]["Question"].unique()
    span = (len(sec_questions)*2)+2
    
    ws.merge_cells(start_row=1,start_column=col_pointer,
                   end_row=1,end_column=col_pointer+span-1)
    ws.cell(row=1,column=col_pointer).value=section
    ws.cell(row=1,column=col_pointer).font=bold
    ws.cell(row=1,column=col_pointer).alignment=center

    for q in sec_questions:
        ws.merge_cells(start_row=2,start_column=col_pointer,
                       end_row=2,end_column=col_pointer+1)
        ws.cell(row=2,column=col_pointer).value=q
        ws.cell(row=2,column=col_pointer).font=bold
        ws.cell(row=2,column=col_pointer).alignment=center
        ws.cell(row=3,column=col_pointer).font=bold
        ws.cell(row=3,column=col_pointer).alignment=center
        ws.cell(row=3,column=col_pointer).value="Percentage"
        ws.cell(row=3,column=col_pointer+1).font=bold
        ws.cell(row=3,column=col_pointer+1).alignment=center
        ws.cell(row=3,column=col_pointer+1).value="Mean"
        col_pointer+=2

    # Section total
    ws.merge_cells(start_row=2,start_column=col_pointer,
                   end_row=2,end_column=col_pointer+1)
    ws.cell(row=2,column=col_pointer).value="Section Total"
    ws.cell(row=2,column=col_pointer).font=bold
    ws.cell(row=2,column=col_pointer).alignment=center
    ws.cell(row=3,column=col_pointer).font=bold
    ws.cell(row=3,column=col_pointer).alignment=center
    ws.cell(row=3,column=col_pointer).value="Percentage"
    ws.cell(row=3,column=col_pointer+1).font=bold
    ws.cell(row=3,column=col_pointer+1).alignment=center
    ws.cell(row=3,column=col_pointer+1).value="Mean"
    col_pointer+=2

# OVERALL HEADER
ws.merge_cells(start_row=1,start_column=col_pointer,
               end_row=1,end_column=col_pointer+1)
ws.cell(row=1,column=col_pointer).font=bold
ws.cell(row=1,column=col_pointer).alignment=center
ws.cell(row=1,column=col_pointer).value="Overall"
ws.merge_cells(start_row=2,start_column=col_pointer,
               end_row=2,end_column=col_pointer+1)
ws.cell(row=2,column=col_pointer).font=bold
ws.cell(row=2,column=col_pointer).alignment=center
ws.cell(row=2,column=col_pointer).value="Overall"
ws.cell(row=3,column=col_pointer).font=bold
ws.cell(row=3,column=col_pointer).alignment=center
ws.cell(row=3,column=col_pointer).value="Percentage"
ws.cell(row=3,column=col_pointer+1).font=bold
ws.cell(row=3,column=col_pointer+1).alignment=center
ws.cell(row=3,column=col_pointer+1).value="Mean"

# WRITE DATA
row_pointer = 4
group_keys = static_cols[:-2]
grouped = agg_df.groupby(group_keys)

for keys, group in grouped:
    col_pointer=1
    for value in keys:
        ws.cell(row_pointer,col_pointer).value=value
        col_pointer+=1
    ws.cell(row=row_pointer,column=col_pointer).alignment=center
    ws.cell(row_pointer,col_pointer).value=group["student_count"].iloc[0]
    col_pointer+=1
    ws.cell(row=row_pointer,column=col_pointer).alignment=center
    ws.cell(row_pointer,col_pointer).value=group["Total_Responses"].max()
    col_pointer+=1

    for section in sections:
        sec_group = group[group["Section Name"]==section]
        for _,q_row in sec_group.iterrows():
            ws.cell(row=row_pointer,column=col_pointer).alignment=center
            ws.cell(row_pointer,col_pointer).value=round(q_row["Percentage"],2)
            ws.cell(row=row_pointer,column=col_pointer+1).alignment=center
            ws.cell(row_pointer,col_pointer+1).value=round(q_row["Mean"],2)
            col_pointer+=2

        sec_total = section_df[
            (section_df["class_id"]==group["class_id"].iloc[0]) &
            (section_df["Section Name"]==section)
        ]
        ws.cell(row=row_pointer,column=col_pointer).alignment=center
        ws.cell(row_pointer,col_pointer).value=round(sec_total["Section_Percentage"].values[0],2)
        ws.cell(row=row_pointer,column=col_pointer+1).alignment=center
        ws.cell(row_pointer,col_pointer+1).value=round(sec_total["Section_Mean"].values[0],2)
        col_pointer+=2

    overall_row = overall_df[
        overall_df["class_id"]==group["class_id"].iloc[0]
    ]
    ws.cell(row=row_pointer,column=col_pointer).alignment=center
    ws.cell(row_pointer,col_pointer).value=round(overall_row["Overall_Percentage"].values[0],2)
    ws.cell(row=row_pointer,column=col_pointer+1).alignment=center
    ws.cell(row_pointer,col_pointer+1).value=round(overall_row["Overall_Mean"].values[0],2)

    row_pointer+=1

# Borders
for row in ws.iter_rows():
    for cell in row:
        cell.border=thin

# SUBJECTIVE SHEET
if not subjective_df.empty:
    ws_sub = wb.create_sheet("Subjective Feedback")
    headers=["Session Name","Class Name","Faculty Name",
             "Course Code","Course Name",
             "Question","Masked Student ID","Response/Option"]
    ws_sub.append(headers)
    for _,r in subjective_df[headers].iterrows():
        ws_sub.append(r.tolist())
    for cell in ws_sub[1]:
        cell.font=bold

# SAVE TO DOWNLOADS
session_name = raw_df["Session Name"].iloc[0]
safe_name = re.sub(r'[\\/*?:"<>|]',"",session_name)
file_name=f"{safe_name}.xlsx"
downloads=Path.home()/"Downloads"/"SGBS Feedback Report"
downloads.mkdir(exist_ok=True)
wb.save(downloads/file_name)

print(f"Report saved to: {downloads/file_name}")
