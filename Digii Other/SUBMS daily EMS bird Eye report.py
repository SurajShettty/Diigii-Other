import pandas as pd
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ============================================================
# DATABASE CONFIG
# ============================================================
DB_CONFIG = {
    'host': 'collpolldb19-read.c5sc77nejhmr.ap-south-1.rds.amazonaws.com',
    'user': 'suraj_shetty',
    'password': 'LW3J0MU3mZ',
    'database': 'collpoll_subms'
}

# ============================================================
# OUTPUT FILE
# ============================================================
today_date = datetime.date.today()
formatted_date = today_date.strftime("%d-%m-%Y %H-%M")
OUTPUT_FILE = rf"C:\Users\suraj\OneDrive\Desktop\SUBMS_EMS Bird Eye View Report with summary and question paper data_{pd.Timestamp.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

# ============================================================
# QUERY
# ============================================================
QUERY = r'''
WITH allocation_summary AS (
    SELECT 
        ea.id AS assessment_id,
        SUM(IF(eavm.student_count IS NULL, 0, eavm.student_count)) AS total_students_allocated,
        COUNT(ea.id) AS total_assessments,
        COUNT(DISTINCT eavm.assessment_venue_infrastructure_id) AS total_venues,
        COUNT(DISTINCT eaim.assessment_venue_infrastructure_id) AS total_invigilator_allocations,
        COUNT(eas.venue_seating_id) AS total_seating_allocations,
        CONCAT(uax.f_name, ' ', uax.m_name, ' ', uax.l_name) AS invigilator_name,
        a.email as invigilator_email,
        uax.registration_id as invigilator_registration_id,
        iv.name as venue
    FROM ems_assessment ea

    LEFT JOIN ems_assessment_venue_mapping eavm
        ON eavm.assessment_id = ea.id

    LEFT JOIN ems_assessment_venue_infrastructure eavii
        ON eavii.id = eavm.assessment_venue_infrastructure_id

    LEFT JOIN infrastructure_version iv
        ON iv.id = eavii.infrastructure_id

    LEFT JOIN ems_assessment_invigilator_mapping eaim
        ON eaim.assessment_venue_infrastructure_id = eavm.assessment_venue_infrastructure_id
        AND ea.slot_date_id = eaim.assessment_slot_date_id

    LEFT JOIN ems_assessment_student eas
        ON eas.assessment_id = ea.id

    LEFT JOIN user_attributes uax
        ON uax.ukid = eaim.faculty_ukid

    LEFT JOIN authenticator a
        ON uax.ukid = a.ukid

    GROUP BY ea.id
)

SELECT
    t.id AS term_id,
    t.name AS term_name,
    tc.id AS term_course_id,
    co.course_code,
    cov.course_name,
    eet.id AS exam_type_id,
    eet.name AS assessment_name,
    easg.group_name AS group_name,

    IF(ea.online = 1, 'online', 'offline') AS exam_mode,

    DATE(ea.start_datetime) AS start_date,
    DAYNAME(ea.start_datetime) AS start_day_name,

    eass.slot_name,

    TIME(ea.start_datetime) AS start_time,
    TIME(ea.closing_datetime) AS end_time,

    ea.duration,

    CASE
        WHEN alloc.total_students_allocated >= alloc.total_assessments
        THEN 'FULLY_ALLOCATED'

        WHEN alloc.total_students_allocated > 0
             AND alloc.total_students_allocated < alloc.total_assessments
        THEN 'PARTIALLY_ALLOCATED'

        ELSE 'NOT_ALLOCATED'
    END AS venue_status,

    alloc.venue,

    CASE
        WHEN alloc.total_venues = alloc.total_invigilator_allocations
             AND alloc.total_invigilator_allocations > 0
        THEN 'FULLY_ALLOCATED'

        WHEN alloc.total_invigilator_allocations > 0
             AND alloc.total_venues > alloc.total_invigilator_allocations
        THEN 'PARTIALLY_ALLOCATED'

        ELSE 'NOT_ALLOCATED'
    END AS invigilator_status,

    alloc.invigilator_name,
    alloc.invigilator_registration_id,
    alloc.invigilator_email,

    CASE
        WHEN alloc.total_assessments = alloc.total_seating_allocations
        THEN 'FULLY_ALLOCATED'

        WHEN alloc.total_seating_allocations > 0
             AND alloc.total_assessments > alloc.total_seating_allocations
        THEN 'PARTIALLY_ALLOCATED'

        ELSE 'NOT_ALLOCATED'
    END AS seating_status

FROM ems_assessment ea

LEFT JOIN allocation_summary alloc
    ON alloc.assessment_id = ea.id

LEFT JOIN term_course tc
    ON tc.id = ea.term_course_id

LEFT JOIN course_version cov
    ON tc.course_version_id = cov.id

LEFT JOIN course co
    ON co.course_id = cov.course_id

LEFT JOIN term t
    ON t.id = tc.term_id

LEFT JOIN ems_examination_type eet
    ON eet.id = ea.exam_type_id

LEFT JOIN ems_assessment_schedule ascc
    ON ascc.id = ea.assessment_schedule_id

LEFT JOIN ems_assessment_schedule_groups easg
    ON easg.id = ea.schedule_group_id

LEFT JOIN ems_assessment_slot_dates easd
    ON easd.id = ea.slot_date_id

LEFT JOIN ems_assessment_schedule_slots eass
    ON eass.id = easd.ems_assessment_slot_id

WHERE t.id = 61
AND DATE(ea.start_datetime) >= '2026-05-07'
AND DATE(ea.start_datetime) <= '2026-05-16'

GROUP BY
    t.id,
    ea.term_course_id,
    co.course_code,
    assessment_name,
    eass.slot_name,
    ea.duration;
'''
# ============================================================
# QUESTION PAPER QUERY
# ============================================================

QUESTION_PAPER_QUERY = r'''
SELECT
    ee.name AS exam_name,
    t.id AS term_id,
    t.name AS term,
    eas.name AS assessment_schedule,
    eet.name AS assessment_type,
    easg.group_name AS exam_group,
    d.department_name,
    cv.course_name,
    c.course_code,
    cv.course_credits,
    eaqps.set_label,
    eaqpse.set_type AS question_paper_type,
    eaqpse.generated_from AS generation_type,
    ea.start_datetime,
    ea.closing_datetime,
    ea.duration,

    IF(ea.online = 1, 'Online', 'Offline') AS exam_type,

    IF(ea.allow_mobile_exam = 1, 'Yes', 'No') AS allowed_mobile_exam,

    IF(ea.shuffle_questions = 1, 'Yes', 'No') AS shuffle_questions,

    IF(ea.shuffle_options = 1, 'Yes', 'No') AS shuffle_options,

    eaqp.status AS question_paper_status,
    eaqps.status AS question_paper_set_status,

    ua.registration_id AS setter_id,

    CONCAT(
        ua.f_name, ' ',
        ua.m_name, ' ',
        ua.l_name
    ) AS setter_name,

    a.email AS setter_email,

    eaqps.status AS setter_status,

    eaqp.submission_deadline AS setter_deadline,

    CONCAT(
        'Reviewer ',
        easr.sequence_number
    ) AS reviewer_number,

    CONCAT(
        ua2.f_name, ' ',
        ua2.m_name, ' ',
        ua2.l_name
    ) AS reviewer_name,

    ua2.registration_id AS reviewer_id,

    a2.email AS reviewer_email,

    easr.status AS reviewer_status

FROM ems_assessment_question_paper eaqp

LEFT JOIN ems_assessment_question_paper_settings eaqpse
    ON eaqpse.term_course_id = eaqp.term_course_id
    AND eaqpse.exam_type_id = eaqp.exam_type_id

LEFT JOIN ems_examination_type eet
    ON eaqp.exam_type_id = eet.id

LEFT JOIN ems_assessment ea
    ON ea.id = eaqp.assessment_id

LEFT JOIN ems_assessment_schedule eas
    ON eas.id = ea.assessment_schedule_id

LEFT JOIN term_course tc
    ON tc.id = eaqp.term_course_id

LEFT JOIN ems_examination ee
    ON ee.term_id = tc.term_id

LEFT JOIN term t
    ON t.id = tc.term_id

LEFT JOIN course c
    ON tc.course_id = c.course_id

LEFT JOIN course_version cv
    ON cv.id = tc.course_version_id

LEFT JOIN department d
    ON c.department_id = d.department_id

LEFT JOIN ems_assessment_schedule_groups easg
    ON ee.id = easg.exam_id
    AND tc.id = easg.term_course_id
    AND eet.id = easg.exam_type_id

LEFT JOIN ems_assessment_question_paper_set eaqps
    ON eaqps.question_paper_id = eaqp.id

LEFT JOIN ems_assessment_question_paper_setter eaqpss
    ON eaqpss.question_paper_id = eaqp.id
    AND eaqps.question_paper_setter_id = eaqpss.id

LEFT JOIN user_attributes ua
    ON ua.ukid = eaqpss.setter_ukid

LEFT JOIN authenticator a
    ON a.ukid = ua.ukid

LEFT JOIN ems_assessment_set_reviewer easr
    ON easr.question_paper_id = eaqp.id
    AND easr.question_paper_set_id = eaqps.id

LEFT JOIN user_attributes ua2
    ON ua2.ukid = easr.reviewer_ukid

LEFT JOIN authenticator a2
    ON a2.ukid = ua2.ukid

WHERE t.id = 61

AND DATE(ea.start_datetime) >= '2026-05-07'
AND DATE(ea.start_datetime) <= '2026-05-16'

ORDER BY easr.reviewer_ukid;
'''
# ============================================================
# CONNECT DATABASE
# ============================================================
conn = mysql.connector.connect(**DB_CONFIG)

# ============================================================
# FETCH DATA
# ============================================================
df = pd.read_sql(QUERY, conn)

question_paper_df = pd.read_sql(
    QUESTION_PAPER_QUERY,
    conn
)

# ============================================================
# FORMAT QUESTION PAPER DURATION
# ============================================================

question_paper_df['duration'] = (
    question_paper_df['duration']
    .astype(str)
    .str.extract(r'(\d{2}:\d{2}:\d{2})', expand=False)
)

conn.close()

# ============================================================
# FORMAT DATE & TIME COLUMNS
# ============================================================

# ============================================================
# FORMAT DATE
# ============================================================

# ============================================================
# FORMAT DATE
# ============================================================

df['start_date'] = pd.to_datetime(
    df['start_date']
).dt.strftime('%d-%m-%Y')

# ============================================================
# FORMAT TIME
# ============================================================

df['start_time'] = (
    df['start_time']
    .astype(str)
    .str.extract(r'(\d{2}:\d{2}:\d{2})', expand=False)
)

df['end_time'] = (
    df['end_time']
    .astype(str)
    .str.extract(r'(\d{2}:\d{2}:\d{2})', expand=False)
)

# ============================================================
# FORMAT DURATION
# ============================================================

df['duration'] = (
    df['duration']
    .astype(str)
    .str.extract(r'(\d{2}:\d{2}:\d{2})', expand=False)
)

# ============================================================
# CLEAN ALL STRING COLUMNS
# ============================================================

for col in df.columns:

    df[col] = df[col].astype(str).str.replace(
        r'[\n\r\t]',
        '',
        regex=True
    )

# ============================================================
# SUMMARY PREP
# ============================================================
all_dates = sorted(df['start_date'].unique())

summary_rows = [
    'EMS Timetable Upload (Courses)',
    'Total No of Mapped Venues',
    'Invigilator Mapping Status',
    'Invigilator not Mapped Status',
    'Seating Plan Generated',
    'Seating Plan not Generated',
    'Question Paper - Not Assigned',
    'Question Paper - Assigned',
    'Question Paper - Submitted',
    'Question Paper - Partially Submitted',
    'Question Paper - Finalized and Published',
    'Online Exam',
    'Offline Exam'
]

summary_df = pd.DataFrame(index=summary_rows)

# ============================================================
# BUILD SUMMARY
# ============================================================
for dt in all_dates:

    temp = df[df['start_date'] == dt]

    summary_df.loc['EMS Timetable Upload (Courses)', dt] = len(temp)

    summary_df.loc['Total No of Mapped Venues', dt] = (
        temp[temp['venue_status'] != 'NOT_ALLOCATED'].shape[0]
    )

    summary_df.loc['Invigilator Mapping Status', dt] = (
        temp[temp['invigilator_status'] == 'FULLY_ALLOCATED'].shape[0]
    )

    summary_df.loc['Invigilator not Mapped Status', dt] = (
        temp[temp['invigilator_status'] == 'NOT_ALLOCATED'].shape[0]
    )

    summary_df.loc['Seating Plan Generated', dt] = (
        temp[temp['seating_status'] == 'FULLY_ALLOCATED'].shape[0]
    )

    summary_df.loc['Seating Plan not Generated', dt] = (
        temp[temp['seating_status'] == 'NOT_ALLOCATED'].shape[0]
    )

    # ========================================================
    # QUESTION PAPER ROWS EMPTY
    # ========================================================
    summary_df.loc['Question Paper - Not Assigned', dt] = ''
    summary_df.loc['Question Paper - Assigned', dt] = ''
    summary_df.loc['Question Paper - Submitted', dt] = ''
    summary_df.loc['Question Paper - Partially Submitted', dt] = ''
    summary_df.loc['Question Paper - Finalized and Published', dt] = ''

    summary_df.loc['Online Exam', dt] = (
        temp[temp['exam_mode'] == 'online'].shape[0]
    )

    summary_df.loc['Offline Exam', dt] = (
        temp[temp['exam_mode'] == 'offline'].shape[0]
    )

# ============================================================
# TOTALS
# ============================================================
summary_df['Pending Courses'] = ''

summary_df['Total Online Exam'] = ''
summary_df.loc['Online Exam', 'Total Online Exam'] = (
    df[df['exam_mode'] == 'online'].shape[0]
)

summary_df['Total Offline Exam'] = ''
summary_df.loc['Offline Exam', 'Total Offline Exam'] = (
    df[df['exam_mode'] == 'offline'].shape[0]
)

# ============================================================
# EXPORT EXCEL
# ============================================================
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:

    # Detailed sheet

    # TAB 1 - MAIN DATA
    # ============================================================

    df.to_excel(
        writer,
        sheet_name='Detailed_Data',
        index=False
    )

    # ============================================================
    # TAB 2 - QUESTION PAPER DATA
    # ============================================================

    question_paper_df.to_excel(
        writer,
        sheet_name='Question_Paper_Data',
        index=False
    )

    # Summary sheet
    summary_export = summary_df.reset_index()
    summary_export.rename(
        columns={'index': 'Examination Details'},
        inplace=True
    )

    summary_export.to_excel(
        writer,
        sheet_name='Summary',
        index=False
    )

# ============================================================
# LOAD WORKBOOK
# ============================================================
wb = load_workbook(OUTPUT_FILE)

ws = wb['Summary']

# ============================================================
# STYLES
# ============================================================
yellow_fill = PatternFill(
    start_color='FFFF00',
    end_color='FFFF00',
    fill_type='solid'
)

bold_font = Font(bold=True)

center_align = Alignment(
    horizontal='center',
    vertical='center'
)

left_align = Alignment(
    horizontal='left',
    vertical='center'
)

thin = Side(style='thin')

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# ============================================================
# APPLY STYLE
# ============================================================
for row in ws.iter_rows():

    for cell in row:

        cell.border = border
        cell.alignment = center_align

# ============================================================
# HEADER STYLE
# ============================================================
for cell in ws[1]:

    cell.fill = yellow_fill
    cell.font = bold_font

# ============================================================
# FIRST COLUMN
# ============================================================
for row in range(2, ws.max_row + 1):

    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).alignment = left_align

# ============================================================
# COLUMN WIDTH
# ============================================================
for column_cells in ws.columns:

    max_length = 0

    column = column_cells[0].column

    for cell in column_cells:

        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    adjusted_width = max_length + 5

    ws.column_dimensions[
        get_column_letter(column)
    ].width = adjusted_width

# ============================================================
# ROW HEIGHT
# ============================================================
for row in range(1, ws.max_row + 1):

    ws.row_dimensions[row].height = 22

# ============================================================
# MERGE TOTAL COLUMNS
# ============================================================

# Pending Courses Column
ws.merge_cells('K2:K14')
ws['K2'] = ''

# Total Online Exam Column
ws.merge_cells('L2:L14')
ws['L2'] = df[df['exam_mode'] == 'online'].shape[0]

# Total Offline Exam Column
ws.merge_cells('M2:M14')
ws['M2'] = df[df['exam_mode'] == 'offline'].shape[0]

# ============================================================
# ALIGNMENT FOR MERGED CELLS
# ============================================================

for cell in ['K2', 'L2', 'M2']:

    ws[cell].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    ws[cell].font = Font(bold=True)



# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT_FILE)

print(f'Excel generated successfully: {OUTPUT_FILE}')