import os
from flask import Flask, render_template, request, send_file, session, redirect, url_for, flash
import openpyxl
from io import BytesIO
from uuid import uuid4

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Required for session

# Initialize data structure to hold all form data
form_data = {
    'servicedesk': [],
    'service': [],
    'form': [],
    'fields': [],
    'option_for_dropdown': [],
    'attachment': [],
    'workcentre': [],
    'workcentre_action': [],
    'workcentre_action_form': [],
    'workcentre_action_form_fields': [],
    'workcentre_action_form_attachment': []
}

def generate_excel(data):
    wb = openpyxl.Workbook()
    
   
    
  

    # Create and populate ServiceDesk sheet
    ws = wb.create_sheet("servicedesk")
    ws.append(["title", "icon", "parentServiceDeskName", "parentServiceDeskId"])
    for item in data['servicedesk']:
        ws.append([item.get('title', ''), item.get('icon', 'default'), 
                  item.get('parentServiceDeskName', ''), item.get('parentServiceDeskId', '')])

    
        
    # Create and populate Service sheet
    ws = wb.create_sheet("service")
    service_headers = ["id", "title", "description", "icon", "serviceDeskName", "serviceDeskId", 
                      "paid", "amount", "currency", "costItemId", "escalationEnabled", 
                      "commentsDisabled", "deleted", "anonymous", "sendEmail", 
                      "duplicateRequest", "isFeedbackEnabled", "workingHoursEnabled", 
                      "isMiddlewareService"]
    ws.append(service_headers)
    for item in data['service']:
        row = [item.get(col, '') for col in service_headers]
        ws.append(row)
    
    # Create and populate Form sheet
    ws = wb.create_sheet("form")
    form_headers = ["title", "description", "autoPopulateEmail", "autoPopulatePhone", 
                   "autoPopulateHostelDetails", "disclaimerRequired", "disclaimer", 
                   "serviceName", "serviceId", "showEmailToResolver", 
                   "showPhoneToResolver", "showHostelDetailsToResolver"]
    ws.append(form_headers)
    for item in data['form']:
        row = [item.get(col, '') for col in form_headers]
        ws.append(row)
    
    # Create and populate Fields sheet
    ws = wb.create_sheet("fields")
    fields_headers = ["id", "name", "element", "minimumLength", "maximumLength", 
                     "minimumNumber", "maximumNumber", "mandatory", "formName", 
                     "formId", "showToResolver", "fieldTypeParent", "userFieldIdentifier"]
    ws.append(fields_headers)
    for item in data['fields']:
        row = [item.get(col, '') for col in fields_headers]
        ws.append(row)
    
    # Create and populate Option for dropdown sheet
    ws = wb.create_sheet("option for dropdown")
    ws.append(["fieldId", "label"])
    for item in data['option_for_dropdown']:
        ws.append([item.get('fieldName', ''), item.get('label', '')])
    
    # Create and populate Attachment sheet
    ws = wb.create_sheet("attachment")
    ws.append(["attachmentLabel", "formName", "formId", "mandatory"])
    for item in data['attachment']:
        ws.append([item.get('attachmentLabel', ''), item.get('formName', ''), 
                item.get('formId', ''), item.get('mandatory', '')])

    
    # Create and populate Workcentre sheet
    ws = wb.create_sheet("workcentre")
    workcentre_headers = ["title", "assignmentType", "reassignmentType", 
                         "parentWorkcentreName", "parentWorkcentreId", 
                         "closureWorkcentre", "serviceName", "serviceId", 
                         "chcAssignmentEntity"]
    ws.append(workcentre_headers)
    for item in data['workcentre']:
        row = [item.get(col, '') for col in workcentre_headers]
        ws.append(row)
    
    # Create and populate Workcentre Action sheet
    ws = wb.create_sheet("workcentre Action")
    action_headers = ["id", "name", "requireNote", "actionTakenName", 
                     "defaultAction", "type", "workcentreName", "workcentreId", 
                     "formAvailable", "terminationAction"]
    ws.append(action_headers)
    for item in data['workcentre_action']:
        row = [item.get(col, '') for col in action_headers]
        ws.append(row)
    
    # Create and populate Workcentre Action Form sheet
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]
    
    # Debug: Print data being processed
    # print("Excel generation - workcentre_action_form data:", data.get('workcentre_action_form', []))
    
    # Create and populate Workcentre Action Form sheet
    if data.get('workcentre_action_form'):
        ws = wb.create_sheet("workcentre_action_form")  # Consistent naming
        headers = ["title", "description", "autoPopulateEmail", 
                  "autoPopulatePhone", "actionId", "showRequester"]
        ws.append(headers)
        
        for item in data['workcentre_action_form']:
            row = [item.get(col, '') for col in headers]
            ws.append(row)
            print("Added row to workcentre_action_form:", row)

    
    # Create and populate Workcentre Action Form Fields sheet
    ws = wb.create_sheet("work centre action form fields")
    action_form_fields_headers = ["name", "element", "minimumLength", "maximumLength", "minimumNumber", "maximumNumber", "mandatory", "formName", "formId"]

    ws.append(action_form_fields_headers)
    for item in data['workcentre_action_form_fields']:
        row = [item.get(col, '') for col in action_form_fields_headers]
        ws.append(row)
    
    # Create and populate Workcentre Action Form Attachment sheet
    ws = wb.create_sheet("work centre action form attachm")
    ws.append(["attachmentLabel", "formName", "formId", "mandatory"])
    for item in data['workcentre_action_form_attachment']:
        ws.append([item.get('attachmentLabel', ''), item.get('formName', ''), 
                  item.get('formId', ''), item.get('mandatory', '')])
    
    
    
    # Save to a BytesIO object
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    return virtual_workbook

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/servicedesk', methods=['GET', 'POST'])
def servicedesk():
    if request.method == 'POST':
        servicedesk_info = {
            'title': request.form.get('title'),
            'icon': request.form.get('icon', 'default'),
            'parentServiceDeskName': request.form.get('parentServiceDeskName', ''),
            'parentServiceDeskId': request.form.get('parentServiceDeskId', '')
        }
        session['servicedesk_info'] = servicedesk_info
        form_data['servicedesk'].append(servicedesk_info)
        return redirect(url_for('service'))
    return render_template('servicedesk.html')

@app.route('/service', methods=['GET', 'POST'])
def service():
    servicedesk_info = session.get('servicedesk_info', {})
    
    if request.method == 'POST':
        service_id = str(uuid4())
        form_data['service'].append({
            'id': 'service_id',
            'title': request.form.get('title'),
            'description': request.form.get('title', ''),
            'icon': request.form.get('icon', 'default'),
            'serviceDeskName': servicedesk_info.get('title', ''),
            'serviceDeskId': servicedesk_info.get('parentServiceDeskId', ''),
            'paid': request.form.get('paid', ''),
            'amount': request.form.get('amount', ''),
            'currency': request.form.get('currency', ''),
            'costItemId': request.form.get('costItemId', ''),
            'escalationEnabled': request.form.get('escalationEnabled', ''),
            'commentsDisabled': request.form.get('commentsDisabled', ''),
            'deleted': request.form.get('deleted', ''),
            'anonymous': request.form.get('anonymous', ''),
            'sendEmail': request.form.get('sendEmail', ''),
            'duplicateRequest': request.form.get('duplicateRequest', ''),
            'isFeedbackEnabled': request.form.get('isFeedbackEnabled', ''),
            'workingHoursEnabled': request.form.get('workingHoursEnabled', ''),
            'isMiddlewareService': request.form.get('isMiddlewareService', 'false')
        })
        session['service_info'] = {
            'title': request.form.get('title'),
            'id': service_id
        }
        return redirect(url_for('form'))
    return render_template('service.html', servicedesk_title=servicedesk_info.get('title', ''))

@app.route('/form', methods=['GET', 'POST'])
def form():
    service_info = session.get('service_info', {})
    
    if request.method == 'POST':
        form_id = str(uuid4())
        form_data['form'].append({
            'title': request.form.get('title'),
            'description': request.form.get('description', ''),
            'autoPopulateEmail': request.form.get('autoPopulateEmail', ''),
            'autoPopulatePhone': request.form.get('autoPopulatePhone', ''),
            'autoPopulateHostelDetails': request.form.get('autoPopulateHostelDetails', ''),
            'disclaimerRequired': request.form.get('disclaimerRequired', 'false'),
            'disclaimer': request.form.get('disclaimer', ''),
            'serviceName': service_info.get('title', ''),
            'serviceId': service_info.get('id', ''),
            'showEmailToResolver': request.form.get('showEmailToResolver', ''),
            'showPhoneToResolver': request.form.get('showPhoneToResolver', ''),
            'showHostelDetailsToResolver': request.form.get('showHostelDetailsToResolver', '')
        })
        session['form_info'] = {
            'title': request.form.get('title'),
            'id': form_id
        }
        return redirect(url_for('fields'))
    return render_template('form.html', service_title=service_info.get('title', ''))

@app.route('/fields', methods=['GET', 'POST'])
def fields():
    form_info = session.get('form_info', {})
    
    if request.method == 'POST':
        # Clear previous fields for this form to avoid duplicates
        form_data['fields'] = [f for f in form_data['fields'] if f.get('formName') != form_info.get('title')]
        form_data['option_for_dropdown'] = [o for o in form_data['option_for_dropdown'] if o.get('formId') != form_info.get('id')]

        # Process all submitted fields
        field_data = []
        i = 0
        while f'field_names_{i}' in request.form:
            field_name = request.form.get(f'field_names_{i}')
            if field_name.strip():  # Only process non-empty fields
                field_id = str(uuid4())
                field_element = request.form.get(f'field_elements_{i}', 'textField')
                
                # Add to fields data
                field_data.append({
                    'id': field_id,
                    'name': field_name,
                    'element': field_element,
                    'minimumLength': request.form.get(f'field_min_length_{i}', ''),
                    'maximumLength': request.form.get(f'field_max_length_{i}', ''),
                    'minimumNumber': request.form.get(f'field_min_number_{i}', ''),
                    'maximumNumber': request.form.get(f'field_max_number_{i}', ''),
                    'mandatory': request.form.get(f'field_mandatory_{i}', '0'),
                    'formName': form_info.get('title', ''),
                    'formId': form_info.get('id', ''),
                    'showToResolver': request.form.get(f'field_show_to_resolver_{i}', '1'),
                    'fieldTypeParent': request.form.get(f'field_parent_{i}', 'null'),
                    'userFieldIdentifier': request.form.get(f'field_identifier_{i}', 'null')
                })

                # Process dropdown options if this is a dropdown field
                if field_element == 'dropdown' and f'field_options_{i}' in request.form:
                    options = request.form.get(f'field_options_{i}').split('\n')
                    for option in options:
                        if option.strip():
                            form_data['option_for_dropdown'].append({
                                'fieldName': field_name,
                                'label': option.strip()
                            })
            i += 1

        # Add all fields at once
        form_data['fields'].extend(field_data)
        return redirect(url_for('attachment'))
    
    return render_template('fields.html', form_title=form_info.get('title', ''))

@app.route('/option_for_dropdown', methods=['GET', 'POST'])
def option_for_dropdown():
    form_info = session.get('form_info', {})
    field_info = session.get('current_field', {})
    
    
    if not field_info or not field_info.get('name'):
        flash("Field name is missing. Please reselect the field.", 'error')
        return redirect(url_for('previous_page'))  # Redirect to where the field is chosen
    
    if request.method == 'POST':
        # Clear old options for this field
        form_data['option_for_dropdown'] = [
            opt for opt in form_data['option_for_dropdown']
            if opt.get('fieldName') != field_info.get('name')
        ]
        
        # Add new options (with fieldName)
        options = request.form.getlist('option')
        for option in options:
            if option.strip():
                form_data['option_for_dropdown'].append({
                    'fieldName': field_info.get('name', 'Unknown Field'),  # Ensure this is set
                    'label': option.strip()
                })
        return redirect(url_for('attachment'))
    
    # Get existing options for this field to display in the form
    existing_options = [
        opt['label'] for opt in form_data['option_for_dropdown']
        if opt.get('fieldName') == field_info.get('name')  # Changed from fieldId to fieldName
    ]
    
    return render_template('option_for_dropdown.html',
                         
                         field_name=field_info.get('name', ''),
                         existing_options=existing_options)

@app.route('/attachment', methods=['GET', 'POST'])
def attachment():
    form_info = session.get('form_info', {})

    if request.method == 'POST':
        if 'form_data' not in session:
            session['form_data'] = {}
        if 'attachment' not in session['form_data']:
            session['form_data']['attachment'] = []

        i = 0
        while f'attachmentLabel_{i}' in request.form:
            label = request.form.get(f'attachmentLabel_{i}', '').strip()
            if label:
                session['form_data']['attachment'].append({
                    'attachmentLabel': label,
                    'formName': form_info.get('title', ''),
                    'formId': form_info.get('id', ''),
                    'mandatory': request.form.get(f'mandatory_{i}', '0')
                })
            i += 1
        session.modified = True
        form_data['attachment'] = session['form_data']['attachment']

        return redirect(url_for('workcentre'))

    return render_template('attachment.html', form_title=form_info.get('title', ''))


@app.route('/workcentre', methods=['GET', 'POST'])
def workcentre():
    service_info = session.get('service_info', {})
    service_name = service_info.get('title', '')
    
    if request.method == 'POST':
        # Initialize form_data in session if not exists
        if 'form_data' not in session:
            session['form_data'] = {'workcentre': [], 'workcentre_action': []}
        
        # Clear previous workcenters for this service
        session['form_data']['workcentre'] = [
            w for w in session['form_data'].get('workcentre', []) 
            if w.get('serviceId') != service_info.get('id')
        ]
        
        temp_workcenters = []
        i = 0
        previous_title = None
        
        while f'workcentre_title_{i}' in request.form:
            title = request.form.get(f'workcentre_title_{i}', '').strip()
            if title:
                full_title = f"{service_name}_{title}"
                
                temp_workcenters.append({
                    'title': full_title,
                    'assignmentType': request.form.get(f'workcentre_assignmentType_{i}', 'defaultAssign'),
                    'reassignmentType': request.form.get(f'workcentre_reassignmentType_{i}', 'manualAssign'),
                    'parentWorkcentreName': previous_title,
                    'parentWorkcentreId': request.form.get(f'workcentre_parentId_{i}', ''),
                    'closureWorkcentre': '0',
                    'serviceName': service_name,
                    'serviceId': service_info.get('id', ''),
                    'chcAssignmentEntity': request.form.get(f'workcentre_entity_{i}', '')
                })
                previous_title = full_title
            i += 1
        
        if temp_workcenters:
            temp_workcenters[-1]['closureWorkcentre'] = '1'
            for j in range(1, len(temp_workcenters)):
                temp_workcenters[j]['parentWorkcentreName'] = temp_workcenters[j-1]['title']
                temp_workcenters[j]['parentWorkcentreId'] = ''
        
        session['form_data']['workcentre'].extend(temp_workcenters)
        session.modified = True  # Ensure session is saved
        return redirect(url_for('workcentre_action'))
    
    return render_template('workcentre.html', 
                         service_name=service_name,
                         existing_workcentres=[
                             w for w in session.get('form_data', {}).get('workcentre', [])
                             if w.get('serviceId') == service_info.get('id')
                         ])

@app.route('/workcentre_action', methods=['GET', 'POST'])
def workcentre_action():
    service_info = session.get('service_info', {})
    service_id = service_info.get('id', '')
    
    if 'form_data' not in session:
        session['form_data'] = {'workcentre': [], 'workcentre_action': []}
    
    if request.method == 'POST':
        action_id = str(uuid4())
        action_type = request.form.get('type', 'positive')
        selected_wc_title = request.form.get('workcentreName', '')
        action_name = request.form.get('name', '').strip()
        form_available = request.form.get('formAvailable', '0')
        
        # Create concatenated name: workcentreName_actionName
        concatenated_name =  action_name
        
        new_action = {
            'id': action_id,
            'name': concatenated_name,
            'requireNote': '',
            'actionTakenName': request.form.get('actionTakenName', '').strip(),
            'defaultAction': '1' if action_type == 'positive' else '0',
            'type': action_type,
            'workcentreName': selected_wc_title,
            'workcentreId': '',
            'formAvailable': form_available,
            'terminationAction': request.form.get('terminationAction', '0'),
            'serviceId': service_id,
            'originalActionName': action_name
        }
        
        if 'form_data' not in session:
            session['form_data'] = {}

        if 'workcentre_action' not in session['form_data']:
            session['form_data']['workcentre_action'] = []

        session['form_data']['workcentre_action'].append(new_action)
        session.modified = True

        session['action_info'] = {
            'id': action_id,
            'name': concatenated_name,
            'original_name': action_name
        }
        session.modified = True
        
        # If form is needed, create it automatically and go to fields
        if form_available == '1':
            # Auto-generate a form entry and add to form_data
            form_title = f"{selected_wc_title}_{action_name}".strip()
            form_id = str(uuid4())
            
            form_entry = {
                'title': form_title,
                'description': '',  # No manual description
                'autoPopulateEmail': '',
                'autoPopulatePhone': '',
                'actionId': action_id,
                'showRequester': '1'
            }

            # Save in global form_data
            form_data['workcentre_action_form'].append(form_entry)

            # Save in session too (optional, for continuity)
            if 'form_data' not in session:
                session['form_data'] = {}
            if 'workcentre_action_form' not in session['form_data']:
                session['form_data']['workcentre_action_form'] = []
            session['form_data']['workcentre_action_form'].append(form_entry)

            # Also store the action form info for next step
            session['action_form_info'] = {
                'title': form_title,
                'id': form_id
            }

            session.modified = True
            return redirect(url_for('workcentre_action_form_fields'))

        
        if 'add_another' in request.form:
            return redirect(url_for('workcentre_action'))
        else:
            return redirect(url_for('review'))  # Where to go if no form needed
    
    service_workcentres = [
        wc for wc in session['form_data'].get('workcentre', [])
        if wc.get('serviceId') == service_id
    ]
    
    existing_actions = [
        action for action in session['form_data'].get('workcentre_action', [])
        if action.get('serviceId') == service_id
    ]
    
    return render_template('workcentre_action.html',
                         workcentre_title=service_info.get('title', ''),
                         workcentres=service_workcentres,
                         existing_actions=existing_actions,
                         service_name=service_info.get('title', ''))

# @app.route('/workcentre_action_form', methods=['POST'])
# def workcentre_action_form():
#     action_info = session.get('action_info', {})
    
#     form_id = str(uuid4())
#     form_title = action_info.get('name', '')
    
#     # Create complete form entry
#     form_entry = {
#         'title': form_title,
#         'description': request.form.get('description', ''),
#         'autoPopulateEmail': request.form.get('autoPopulateEmail', '0'),
#         'autoPopulatePhone': request.form.get('autoPopulatePhone', '0'),
#         'actionId': action_info.get('id', ''),
#         'showRequester': request.form.get('showRequester', '0')
#     }
    
#     # Store in both global form_data and session
#     if 'workcentre_action_form' not in form_data:
#         form_data['workcentre_action_form'] = []
#     form_data['workcentre_action_form'].append(form_entry)
    
#     # Initialize session form_data if not exists
#     if 'form_data' not in session:
#         session['form_data'] = {}

#     if 'workcentre_action_form' not in session['form_data']:
#         session['form_data']['workcentre_action_form'] = []

#     session['form_data']['workcentre_action_form'].append(form_entry)
#     session.modified = True
    

#     return redirect(url_for('workcentre_action_form_fields'))
    
    

@app.route('/workcentre_action_form_fields', methods=['GET', 'POST'])
def workcentre_action_form_fields():
    action_form_info = session.get('action_form_info', {})
    
    # Get all actions with formAvailable = '1' (yes)
    form_available_actions = []
    if 'form_data' in session and 'workcentre_action' in session['form_data']:
        form_available_actions = [
            action for action in session['form_data']['workcentre_action']
            if action.get('formAvailable') == '1'
        ]
    
    if request.method == 'POST':
        names = request.form.getlist('name[]')
        elements = request.form.getlist('element[]')
        min_lengths = request.form.getlist('minimum_length[]')
        max_lengths = request.form.getlist('maximum_length[]')
        min_numbers = request.form.getlist('minimum_number[]')
        max_numbers = request.form.getlist('maximum_number[]')
        mandatories = request.form.getlist('mandatory[]')
        inherited = request.form.get('inherited_form', '')
        form_fields = session.setdefault('form_data', {}).setdefault('workcentre_action_form_fields', [])

        # Also update global form_data so generate_excel works
        form_data['workcentre_action_form_fields'] = form_fields

        for i in range(len(names)):
            field = {
                'name': names[i],
                'element': elements[i],
                # 'minimum length': min_lengths[i],
                # 'maximum length': max_lengths[i],
                # 'minimum number': min_numbers[i],
                # 'maximum number': max_numbers[i],
                'minimum length': '',
                'maximum length': '',
                'minimum number': '',
                'maximum number': '',
                'mandatory': '1' if str(i) in mandatories else '0',
                'formName': action_form_info.get('title', ''),
                'formId': action_form_info.get('id', ''),
                'inheritedFrom': inherited
            }
            form_fields.append(field)

        return redirect(url_for('workcentre_action_form_attachment'))

    
    return render_template('workcentre_action_form_fields.html', 
                         form_title=action_form_info.get('title', ''),
                         available_forms=form_available_actions)

@app.route('/workcentre_action_form_attachment', methods=['GET', 'POST'])
def workcentre_action_form_attachment():
    action_form_info = session.get('action_form_info', {})

    # Get all actions with formAvailable = '1'
    form_available_actions = []
    if 'form_data' in session and 'workcentre_action' in session['form_data']:
        form_available_actions = [
            action for action in session['form_data']['workcentre_action']
            if action.get('formAvailable') == '1'
        ]

    if request.method == 'POST':
        # Ensure the list exists
        if 'form_data' not in session:
            session['form_data'] = {}
        if 'workcentre_action_form_attachment' not in session['form_data']:
            session['form_data']['workcentre_action_form_attachment'] = []

        # Get all attachmentLabel fields
        attachment_labels = request.form.getlist('attachmentLabel')
        inherited_forms = request.form.getlist('inherited_form')
        mandatory_flags = request.form.getlist('mandatory')

        # Loop through inputs
        for i in range(len(attachment_labels)):
            label = attachment_labels[i].strip()
            inherited = inherited_forms[i].strip() if i < len(inherited_forms) else ''
            mandatory = mandatory_flags[i] if i < len(mandatory_flags) else '0'

        if label:  # Only save if label is provided
                session['form_data']['workcentre_action_form_attachment'].append({
                    'attachmentLabel': label,
                    'formName': action_form_info.get('title', ''),
                    'formId': action_form_info.get('id', ''),
                    'mandatory': mandatory,
                    'inheritedFrom': inherited
                })

        session.modified = True

        if 'add_another' in request.form:
            return redirect(url_for('workcentre_action'))
        else:
            return redirect(url_for('review'))

    return render_template(
        'workcentre_action_form_attachment.html',
        form_title=action_form_info.get('title', ''),
        available_forms=form_available_actions
    )



@app.route('/review')
def review():
    # Merge session data with form_data
    if 'form_data' in session:
        form_data['workcentre'] = session['form_data'].get('workcentre', [])
        form_data['workcentre_action'] = session['form_data'].get('workcentre_action', [])
    return render_template('review.html', data=form_data)

@app.route('/generate', methods=['POST'])
def generate():
    try:
        # Merge all session data with form_data
        if 'form_data' in session:
            for category in ['workcentre', 'workcentre_action', 'workcentre_action_form',
                             'workcentre_action_form_fields', 'workcentre_action_form_attachment',
                             'attachment']:
                if category in session['form_data']:
                    # Replace existing data with session data to avoid duplicates
                    form_data[category] = session['form_data'][category]

        # 🛠️ Fix: merge attachment data from session if not already merged
        if 'attachment' in session.get('form_data', {}):
            form_data['attachment'] = session['form_data']['attachment']

        virtual_workbook = generate_excel(form_data)
        return send_file(
            virtual_workbook,
            as_attachment=True,
            download_name='CHC_Configuration.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error during generation: {str(e)}")
        flash(f"Error generating Excel: {str(e)}", 'error')
        return redirect(url_for('review'))


if __name__ == '__main__':
    app.run(debug=True)