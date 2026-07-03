"""
Generate pivoted marksheets from a flat grade dataset.

Input : one row per (student, course) with columns
        registration_id, name, programme_name, batch, course_name,
        course_code, course_credits, grade, grade_point, sgpa, cgpa

Output: one .xlsx file per programme-batch, named <programme>_<batch>.xlsx
        Each file contains a single sheet with:
          1. A pivot table  -> rows = students, columns = course_code,
             values = obtained grade (letter), plus trailing SGPA / CGPA columns.
          2. Below it, a course summary (course_code, course_name, credits).

Usage:  python generate_marksheets.py <input_file> [output_dir]
        input_file may be .csv or .xlsx
"""

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- output location (edit this to change where files are written) -------
OUTPUT_DIR = r"C:\Users\suraj\OneDrive\Desktop\marksheets"

# ---- column names in the source file -------------------------------------
COL_REG      = "registration_id"
COL_NAME     = "name"
COL_PROG     = "programme_name"
COL_BATCH    = "batch"
COL_CNAME    = "course_name"
COL_CCODE    = "course_code"
COL_CREDITS  = "course_credits"
COL_GRADE    = "grade"
COL_GPOINT   = "grade_point"
COL_SGPA     = "sgpa"
COL_CGPA     = "cgpa"

# ---- styling -------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
SECTION_FONT  = Font(bold=True, size=12, color="1F4E78")
SUB_FILL      = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT      = Font(bold=True)
THIN          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left", vertical="center")


def sanitize_filename(text: str) -> str:
    """Make a string safe for a Windows filename."""
    text = str(text).strip()
    text = re.sub(r'[<>:"/\\|?*]', "", text)   # strip illegal chars
    text = re.sub(r"\s+", "_", text)            # spaces -> underscores
    return text.strip("._") or "sheet"


def load_data(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(path, dtype=str)
    else:
        # try common encodings; Windows exports are usually cp1252, not utf-8
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                df = pd.read_csv(path, dtype=str, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            df = pd.read_csv(path, dtype=str, encoding="latin-1")
    df.columns = [c.strip() for c in df.columns]
    # normalise whitespace on the text keys we group / display on
    for c in (COL_REG, COL_NAME, COL_PROG, COL_BATCH, COL_CNAME, COL_CCODE, COL_GRADE):
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    # dtype=str turns empty cells into the literal "nan"/"none" -> blank them
    df = df.replace(
        to_replace=r"(?i)^\s*(nan|none|nat|null)\s*$", value="", regex=True
    )
    return df


def first_present(series: pd.Series) -> str:
    """First non-empty value in a series (used for sgpa/cgpa/name)."""
    for v in series:
        if v is not None and str(v).strip() not in ("", "nan", "None"):
            return str(v).strip()
    return ""


def round2(value: str):
    """Round a numeric value to 2 decimals; leave non-numeric text untouched."""
    if value is None or str(value).strip() in ("", "nan", "None"):
        return ""
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return value


def build_group_sheet(ws, group: pd.DataFrame) -> None:
    # ---- course ordering: keep first-seen order ---------------------------
    course_order = list(dict.fromkeys(group[COL_CCODE].tolist()))

    # course_code -> (course_name, credits) for the summary block
    course_info = {}
    for code in course_order:
        rows = group[group[COL_CCODE] == code]
        course_info[code] = (
            first_present(rows[COL_CNAME]) if COL_CNAME in group else "",
            first_present(rows[COL_CREDITS]) if COL_CREDITS in group else "",
        )

    # ---- pivot: student x course -> grade ---------------------------------
    pivot = group.pivot_table(
        index=[COL_REG, COL_NAME],
        columns=COL_CCODE,
        values=COL_GRADE,
        aggfunc="first",
    )
    pivot = pivot.reindex(columns=course_order)  # preserve order
    pivot = pivot.reset_index()

    # per-student sgpa / cgpa carried from the source file
    sgpa_map, cgpa_map = {}, {}
    for reg, rows in group.groupby(COL_REG):
        sgpa_map[reg] = round2(first_present(rows[COL_SGPA])) if COL_SGPA in group else ""
        cgpa_map[reg] = round2(first_present(rows[COL_CGPA])) if COL_CGPA in group else ""

    # ---- write header row -------------------------------------------------
    headers = ["Sl. No", "Reg. No", "Name"] + course_order + ["SGPA", "CGPA"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = LEFT if col_idx in (2, 3) else CENTER
        cell.border = BORDER

    # ---- write student rows ----------------------------------------------
    for serial, (_, prow) in enumerate(pivot.iterrows(), start=1):
        reg = prow[COL_REG]
        row_vals = [serial, reg, prow[COL_NAME]]
        for code in course_order:
            val = prow.get(code)
            row_vals.append("" if pd.isna(val) else val)
        row_vals.append(sgpa_map.get(reg, ""))
        row_vals.append(cgpa_map.get(reg, ""))
        ws.append(row_vals)
        r = ws.max_row
        sgpa_col, cgpa_col = len(row_vals) - 1, len(row_vals)
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER
            cell.alignment = LEFT if col_idx in (2, 3) else CENTER
            if col_idx in (sgpa_col, cgpa_col) and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    pivot_end_row = ws.max_row

    # ---- course summary block (below the pivot) --------------------------
    start = pivot_end_row + 3
    title = ws.cell(row=start, column=1, value="Course Summary")
    title.font = SECTION_FONT

    sub_hdr = ["Sl. No", "Course Code", "Course Name", "Course Credits"]
    for j, h in enumerate(sub_hdr, start=1):
        cell = ws.cell(row=start + 1, column=j, value=h)
        cell.fill = SUB_FILL
        cell.font = SUB_FONT
        cell.border = BORDER
        cell.alignment = LEFT if j == 3 else CENTER

    for i, code in enumerate(course_order, start=1):
        cname, credits = course_info[code]
        sl_cell = ws.cell(row=start + 1 + i, column=1, value=i)
        sl_cell.border = BORDER
        sl_cell.alignment = CENTER
        ws.cell(row=start + 1 + i, column=2, value=code).border = BORDER
        ws.cell(row=start + 1 + i, column=3, value=cname).border = BORDER
        credit_cell = ws.cell(row=start + 1 + i, column=4, value=credits)
        credit_cell.border = BORDER
        credit_cell.alignment = CENTER

    # ---- column widths ----------------------------------------------------
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    ws.freeze_panes = "D2"


def main() -> None:
    if len(sys.argv) >= 2:
        in_path = Path(sys.argv[1])
    else:
        raw = input("Enter the path to the input file (.csv or .xlsx): ").strip().strip('"')
        in_path = Path(raw)

    if not in_path.exists():
        print(f"Input file not found: {in_path}")
        sys.exit(1)

    out_dir = Path(OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    df = load_data(in_path)

    required = [COL_REG, COL_NAME, COL_PROG, COL_BATCH, COL_CCODE, COL_GRADE]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"Missing required columns: {missing}")
        print(f"Found columns: {list(df.columns)}")
        sys.exit(1)

    written = 0
    for (prog, batch), group in df.groupby([COL_PROG, COL_BATCH], sort=True):
        wb = Workbook()
        ws = wb.active
        ws.title = "Marksheet"
        build_group_sheet(ws, group.copy())

        fname = f"{sanitize_filename(prog)}_{sanitize_filename(batch)}.xlsx"
        wb.save(out_dir / fname)
        written += 1
        print(f"  wrote {fname}  ({group[COL_REG].nunique()} students, "
              f"{group[COL_CCODE].nunique()} courses)")

    print(f"\nDone. {written} file(s) written to: {out_dir}")


if __name__ == "__main__":
    main()
