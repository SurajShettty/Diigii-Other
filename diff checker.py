import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

FILE1 = r"C:\Users\suraj\OneDrive\Desktop\ZDA24B039.csv"
FILE2 = r"C:\Users\suraj\OneDrive\Desktop\ZDA24B039 - Copy.csv"
OUTPUT_FILE = r"C:\Users\suraj\OneDrive\Desktop\comparison_report.xlsx"

IGNORE_COLUMNS = ["created_timestamp"]

# ---------------------------
# Read Files
# ---------------------------

def read_file(file):
    if file.lower().endswith(".csv"):
        return pd.read_csv(file, dtype=str)
    return pd.read_excel(file, dtype=str)

df1 = read_file(FILE1).fillna("")
df2 = read_file(FILE2).fillna("")

# Remove ignored columns
df1 = df1.drop(columns=IGNORE_COLUMNS, errors="ignore")
df2 = df2.drop(columns=IGNORE_COLUMNS, errors="ignore")

# Make columns identical
all_cols = list(set(df1.columns).union(set(df2.columns)))

df1 = df1.reindex(columns=all_cols, fill_value="")
df2 = df2.reindex(columns=all_cols, fill_value="")

max_rows = max(len(df1), len(df2))

df1 = df1.reindex(range(max_rows), fill_value="")
df2 = df2.reindex(range(max_rows), fill_value="")

output_df = df2.copy()

modified_cells = []
added_rows = []
deleted_rows = []

# ---------------------------
# Compare
# ---------------------------

for row in range(max_rows):

    row1_empty = (df1.iloc[row] == "").all()
    row2_empty = (df2.iloc[row] == "").all()

    if row1_empty and not row2_empty:
        added_rows.append(row)

    elif not row1_empty and row2_empty:
        deleted_rows.append(row)
        output_df.iloc[row] = df1.iloc[row]

    else:
        for col in all_cols:

            val1 = str(df1.at[row, col])
            val2 = str(df2.at[row, col])

            if val1 != val2:

                output_df.at[row, col] = f"{val1} → {val2}"
                modified_cells.append((row, col))

# ---------------------------
# Save Excel
# ---------------------------

output_df.to_excel(
    OUTPUT_FILE,
    sheet_name="Comparison",
    index=False
)

# ---------------------------
# Highlighting
# ---------------------------

wb = load_workbook(OUTPUT_FILE)
ws = wb["Comparison"]

yellow = PatternFill(
    fill_type="solid",
    start_color="FFEB9C"
)

green = PatternFill(
    fill_type="solid",
    start_color="C6EFCE"
)

red = PatternFill(
    fill_type="solid",
    start_color="FFC7CE"
)

col_map = {
    col: idx + 1
    for idx, col in enumerate(output_df.columns)
}

# Modified cells
for row, col in modified_cells:
    excel_row = row + 2
    excel_col = col_map[col]

    ws.cell(
        excel_row,
        excel_col
    ).fill = yellow

# Added rows
for row in added_rows:
    excel_row = row + 2

    for col in range(1, ws.max_column + 1):
        ws.cell(excel_row, col).fill = green

# Deleted rows
for row in deleted_rows:
    excel_row = row + 2

    for col in range(1, ws.max_column + 1):
        ws.cell(excel_row, col).fill = red

ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)

print(f"Generated: {OUTPUT_FILE}")