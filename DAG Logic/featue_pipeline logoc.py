"""
Feature Scheduling Logic
=========================
Figures out which features SHOULD be scheduled for each tenant, based on:
  1. feature_dag_tags  -> which tag(s)/flag(s) each feature depends on
  2. features (flags)  -> which flags are OPEN/CLOSED per tenant
  3. tenant_feature     -> which features are CURRENTLY scheduled per tenant

Rule:
  - Feature tagged MODULE_AGNOSTIC          -> always required (no flag needed)
  - Feature has no tags / tag not a known flag -> UNMAPPED (can't decide automatically)
  - Otherwise -> required only if ANY of its tags is an OPEN flag for that tenant

Output: a single Excel file with 5 sheets:
  Summary, Missing_Schedules, Extra_Schedules_Review, Unmapped_Features, Full_Detail

Usage:
  Just edit the paths in the CONFIG section below and run:
      python schedule_feature_logic.py
"""

import re

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG - edit these paths before running
# ---------------------------------------------------------------------------

TENANT_FEATURE_PATH = "tenant_feature.csv"        # tenant_id, feature_name
FEATURES_FLAGS_PATH = "features.xlsx"             # instance, feature, status (open/closed) - .xlsx or .csv
FEATURE_DAG_TAGS_PATH = "feature_dag_tags.csv"     # feature_name, tags, dag_file
OUTPUT_PATH = "feature_scheduling_report.xlsx"     # output report

ALWAYS_ON_TAG = "MODULE_AGNOSTIC"


# ---------------------------------------------------------------------------
# 1. LOADERS
# ---------------------------------------------------------------------------

def split_tags(raw: str):
    """Split a tag string into clean individual tags.
    Handles both ';' and ',' as separators (source data has an occasional
    typo using a comma instead of a semicolon)."""
    if not raw or pd.isna(raw):
        return []
    parts = re.split(r"[;,]", str(raw))
    return [p.strip() for p in parts if p.strip()]


def load_feature_tags(path: str):
    """feature_dag_tags file -> {feature_name: [tags]}, {feature_name: dag_file}"""
    df = pd.read_csv(path, dtype=str).fillna("")
    feature_tags = {}
    feature_dagfile = {}
    for _, row in df.iterrows():
        fname = row["feature_name"].strip()
        feature_tags[fname] = split_tags(row["tags"])
        feature_dagfile[fname] = row.get("dag_file", "").strip()
    return feature_tags, feature_dagfile


def load_tenant_flags(path: str):
    """features (flags) file -> {tenant: set(open flags)}, set(all known flags)
    Expected columns: instance, feature, ..., status (open/closed)"""
    if path.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path, dtype=str)
    else:
        df = pd.read_csv(path, dtype=str)
    df = df.fillna("")

    tenant_open_flags = {}
    all_known_flags = set()
    for _, row in df.iterrows():
        tenant = str(row["instance"]).strip()
        flag = str(row["feature"]).strip()
        status = str(row["status"]).strip().lower()
        if not tenant or not flag:
            continue
        all_known_flags.add(flag)
        if status == "open":
            tenant_open_flags.setdefault(tenant, set()).add(flag)
    return tenant_open_flags, all_known_flags


def load_tenant_scheduled(path: str):
    """tenant_feature file -> {tenant: set(scheduled feature_name)}"""
    df = pd.read_csv(path, dtype=str).fillna("")
    tenant_scheduled = {}
    for _, row in df.iterrows():
        t = row["tenant_id"].strip()
        fn = row["feature_name"].strip()
        tenant_scheduled.setdefault(t, set()).add(fn)
    return tenant_scheduled


# ---------------------------------------------------------------------------
# 2. CORE LOGIC
# ---------------------------------------------------------------------------

def build_report(feature_tags, feature_dagfile, tenant_open_flags,
                  all_known_flags, tenant_scheduled):
    all_tenants = sorted(set(tenant_open_flags) | set(tenant_scheduled))

    rows = []
    for tenant in all_tenants:
        open_flags = tenant_open_flags.get(tenant, set())
        scheduled = tenant_scheduled.get(tenant, set())

        for feature, tags in feature_tags.items():
            currently_scheduled = feature in scheduled

            if not tags:
                dep_type = "Unknown (unmapped tag)"
                status = "UNMAPPED"
                matched = ""
                reason = "No tags/dag mapping found for this feature"

            elif ALWAYS_ON_TAG in tags:
                dep_type = "Always-on (no flag)"
                status = "REQUIRED"
                matched = ALWAYS_ON_TAG
                reason = "Module-agnostic - always scheduled, no flag dependency"

            else:
                known_tags = [t for t in tags if t in all_known_flags]
                if not known_tags:
                    dep_type = "Unknown (unmapped tag)"
                    status = "UNMAPPED"
                    matched = ""
                    reason = f"Tag(s) [{', '.join(tags)}] not found among known flags"
                else:
                    matched_open = [t for t in known_tags if t in open_flags]
                    dep_type = "Flag-gated"
                    if matched_open:
                        status = "REQUIRED"
                        matched = ", ".join(matched_open)
                        reason = "Matching feature flag is OPEN for this tenant"
                    else:
                        status = "NOT_REQUIRED"
                        matched = ""
                        reason = f"None of tag(s) [{', '.join(known_tags)}] are open"

            if status == "REQUIRED":
                action = "OK - already scheduled" if currently_scheduled else "SCHEDULE (missing)"
            elif status == "NOT_REQUIRED":
                action = "REVIEW - remove from schedule" if currently_scheduled else "OK - correctly not scheduled"
            else:
                action = "REVIEW - unmapped, verify manually"

            rows.append({
                "tenant_id": tenant,
                "feature_name": feature,
                "dependency_type": dep_type,
                "dag_file": feature_dagfile.get(feature, ""),
                "tags": "; ".join(tags),
                "currently_scheduled": "Yes" if currently_scheduled else "No",
                "required_by_flags": status,
                "matched_open_flag": matched,
                "action": action,
                "reason": reason,
            })

    return pd.DataFrame(rows)


def slice_report(full: pd.DataFrame):
    missing = full[full["action"] == "SCHEDULE (missing)"].drop(columns=["reason"]).reset_index(drop=True)
    to_remove = full[full["action"] == "REVIEW - remove from schedule"].drop(columns=["reason"]).reset_index(drop=True)
    unmapped = (
        full[full["required_by_flags"] == "UNMAPPED"]
        [["tenant_id", "feature_name", "dependency_type", "dag_file", "tags", "currently_scheduled", "reason"]]
        .drop_duplicates(subset=["feature_name", "reason"])
        .reset_index(drop=True)
    )

    summary = pd.DataFrame([
        ["Total tenants analyzed", full["tenant_id"].nunique()],
        ["Total features analyzed", full["feature_name"].nunique()],
        ["Total (tenant, feature) combinations", len(full)],
        ["Missing schedules (need to add)", len(missing)],
        ["  - Always-on (no flag needed)", (missing["dependency_type"] == "Always-on (no flag)").sum()],
        ["  - Flag-gated (open flag exists)", (missing["dependency_type"] == "Flag-gated").sum()],
        ["Extra schedules to review (flag closed)", len(to_remove)],
        ["Unmapped feature rows (no tag / unknown tag)", (full["required_by_flags"] == "UNMAPPED").sum()],
        ["Already correctly scheduled", (full["action"] == "OK - already scheduled").sum()],
        ["Already correctly NOT scheduled", (full["action"] == "OK - correctly not scheduled").sum()],
    ], columns=["Metric", "Value"])

    return missing, to_remove, unmapped, summary


# ---------------------------------------------------------------------------
# 3. EXCEL OUTPUT
# ---------------------------------------------------------------------------

def write_excel(output_path, summary, missing, to_remove, unmapped, full):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        missing.to_excel(writer, sheet_name="Missing_Schedules", index=False)
        to_remove.to_excel(writer, sheet_name="Extra_Schedules_Review", index=False)
        unmapped.to_excel(writer, sheet_name="Unmapped_Features", index=False)
        full.to_excel(writer, sheet_name="Full_Detail", index=False)

    wb = openpyxl.load_workbook(output_path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col, max_row = ws.max_column, ws.max_row

        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        if max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        widths = {}
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 2000)):
            for cell in row:
                if cell.row > 1:
                    cell.font = body_font
                length = len(str(cell.value)) if cell.value is not None else 0
                widths[cell.column] = max(widths.get(cell.column, 0), length)
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w + 2, 10), 55)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# 4. MAIN
# ---------------------------------------------------------------------------

def main():
    feature_tags, feature_dagfile = load_feature_tags(FEATURE_DAG_TAGS_PATH)
    tenant_open_flags, all_known_flags = load_tenant_flags(FEATURES_FLAGS_PATH)
    tenant_scheduled = load_tenant_scheduled(TENANT_FEATURE_PATH)

    full = build_report(feature_tags, feature_dagfile, tenant_open_flags,
                         all_known_flags, tenant_scheduled)
    missing, to_remove, unmapped, summary = slice_report(full)

    write_excel(OUTPUT_PATH, summary, missing, to_remove, unmapped, full)
    print(f"Report written to {OUTPUT_PATH}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()