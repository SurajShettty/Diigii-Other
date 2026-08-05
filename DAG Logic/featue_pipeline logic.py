"""
Feature Scheduling Logic
=========================
Figures out which pipelines/features SHOULD be scheduled for each tenant, based on:
  1. pipelines.json    -> per (tenant, dag_id): which feature_flags it depends on,
                          and whether it's currently enabled (enabled: true/false)
  2. features (flags)  -> which flags are OPEN/CLOSED per tenant

Rule:
  - dag tagged MODULE_AGNOSTIC              -> always required (no flag needed)
  - dag has no feature_flags / unknown flag -> UNMAPPED (can't decide automatically)
  - Otherwise -> required only if ANY of its feature_flags is an OPEN flag for that tenant

"Currently scheduled" = enabled: true in pipelines.json.
enabled: false means it was scheduled once and has since been disabled -> NOT currently scheduled.

Output: a single Excel file with 5 sheets:
  Summary, Missing_Schedules, Extra_Schedules_Review, Unmapped_Features, Full_Detail

Usage:
  Just edit the paths in the CONFIG section below and run:
      python schedule_feature_logic.py

Query for features file (run across all instances):
select f.feature,default_name,feature_custom_name,description,status from college_feature_flag t1 left join feature f on f.id = t1.feature_id
"""

import json
import re

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG - edit these paths before running
# ---------------------------------------------------------------------------

PIPELINES_JSON_PATH = "C:\\Users\\suraj\\OneDrive\\Desktop\\pipelines.json"            # tenant_id, dag_id, feature_flags, enabled
FEATURES_FLAGS_PATH = "C:\\Users\\suraj\\Downloads\\features.xlsx"             # instance, feature, status (open/closed) - .xlsx or .csv
OUTPUT_PATH = "C:\\Users\\suraj\\OneDrive\\Desktop\\feature_scheduling_report.xlsx"      # output report

ALWAYS_ON_TAG = "MODULE_AGNOSTIC"


# ---------------------------------------------------------------------------
# 1. LOADERS
# ---------------------------------------------------------------------------

def split_flag_values(raw):
    """Split a single feature_flags entry into clean individual flags.
    Handles both ';' and ',' as separators, in case a single list item
    ever contains multiple comma-separated flags (e.g. "A,B")."""
    if raw is None:
        return []
    parts = re.split(r"[;,]", str(raw))
    return [p.strip() for p in parts if p.strip()]


def load_pipelines_json(path: str):
    """pipelines.json -> 
       dag_tags:        {dag_id: [feature_flags]}
       dag_tables:      {dag_id: [tables]}   (kept for reference in the report)
       tenant_scheduled: {tenant_id: set(dag_id)}  where enabled == True
       all_tenants:     set of every tenant_id seen (enabled true or false)
    """
    with open(path) as f:
        payload = json.load(f)

    # supports either a bare list, or the {"results": [...]} wrapper seen in this file
    records = payload["results"] if isinstance(payload, dict) and "results" in payload else payload

    dag_tags = {}
    dag_tables = {}
    tenant_scheduled = {}
    all_tenants = set()

    for rec in records:
        tenant = str(rec.get("tenant_id", "")).strip()
        dag_id = str(rec.get("dag_id", "")).strip()
        if not tenant or not dag_id:
            continue
        all_tenants.add(tenant)

        # flatten feature_flags, splitting any comma/semicolon-joined entries
        raw_flags = rec.get("feature_flags") or []
        flags = []
        for item in raw_flags:
            flags.extend(split_flag_values(item))
        # keep the union of flags seen for this dag_id (should be consistent
        # across tenants, but union protects against any inconsistency)
        existing = dag_tags.setdefault(dag_id, [])
        for flag in flags:
            if flag not in existing:
                existing.append(flag)

        dag_tables.setdefault(dag_id, rec.get("tables") or [])

        if rec.get("enabled") is True:
            tenant_scheduled.setdefault(tenant, set()).add(dag_id)

    return dag_tags, dag_tables, tenant_scheduled, all_tenants


def load_tenant_flags(path: str):
    """features (flags) file -> {tenant: set(open flags)}, set(all known flags)
    Expected columns: instance, feature, ..., status (open/closed)"""
    if path.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path, dtype=str)
    else:
        df = pd.read_csv(path, dtype=str)
    df = df.fillna("")

    tenant_open_flags = {}
    all_known_flags = set()
    for _, row in df.iterrows():
        tenant = str(row["instance"]).strip()
        flag = str(row["feature"]).strip()
        status = str(row["status"]).strip().lower()
        if not tenant or not flag:
            continue
        all_known_flags.add(flag)
        if status == "open":
            tenant_open_flags.setdefault(tenant, set()).add(flag)
    return tenant_open_flags, all_known_flags


# ---------------------------------------------------------------------------
# 2. CORE LOGIC
# ---------------------------------------------------------------------------

def build_report(dag_tags, dag_tables, tenant_open_flags,
                  all_known_flags, tenant_scheduled, all_tenants):
    all_tenants = sorted(all_tenants | set(tenant_open_flags) | set(tenant_scheduled))

    rows = []
    for tenant in all_tenants:
        open_flags = tenant_open_flags.get(tenant, set())
        scheduled = tenant_scheduled.get(tenant, set())

        for dag_id, tags in dag_tags.items():
            currently_scheduled = dag_id in scheduled

            if not tags:
                dep_type = "Unknown (unmapped tag)"
                status = "UNMAPPED"
                matched = ""
                reason = "No feature_flags found for this pipeline"

            elif ALWAYS_ON_TAG in tags:
                dep_type = "Always-on (no flag)"
                status = "REQUIRED"
                matched = ALWAYS_ON_TAG
                reason = "Module-agnostic - always scheduled, no flag dependency"

            else:
                known_tags = [t for t in tags if t in all_known_flags]
                if not known_tags:
                    dep_type = "Unknown (unmapped tag)"
                    status = "UNMAPPED"
                    matched = ""
                    reason = f"Flag(s) [{', '.join(tags)}] not found among known flags"
                else:
                    matched_open = [t for t in known_tags if t in open_flags]
                    dep_type = "Flag-gated"
                    if matched_open:
                        status = "REQUIRED"
                        matched = ", ".join(matched_open)
                        reason = "Matching feature flag is OPEN for this tenant"
                    else:
                        status = "NOT_REQUIRED"
                        matched = ""
                        reason = f"None of flag(s) [{', '.join(known_tags)}] are open"

            if status == "REQUIRED":
                action = "OK - already scheduled" if currently_scheduled else "SCHEDULE (missing)"
            elif status == "NOT_REQUIRED":
                action = "REVIEW - remove from schedule" if currently_scheduled else "OK - correctly not scheduled"
            else:
                action = "REVIEW - unmapped, verify manually"

            rows.append({
                "tenant_id": tenant,
                "dag_id": dag_id,
                "dependency_type": dep_type,
                "tables": ", ".join(dag_tables.get(dag_id, [])),
                "feature_flags": "; ".join(tags),
                "currently_scheduled": "Yes" if currently_scheduled else "No",
                "required_by_flags": status,
                "matched_open_flag": matched,
                "action": action,
                "reason": reason,
            })

    return pd.DataFrame(rows)


def slice_report(full: pd.DataFrame):
    missing = full[full["action"] == "SCHEDULE (missing)"].drop(columns=["reason"]).reset_index(drop=True)
    to_remove = full[full["action"] == "REVIEW - remove from schedule"].drop(columns=["reason"]).reset_index(drop=True)
    unmapped = (
        full[full["required_by_flags"] == "UNMAPPED"]
        [["tenant_id", "dag_id", "dependency_type", "tables", "feature_flags", "currently_scheduled", "reason"]]
        .drop_duplicates(subset=["dag_id", "reason"])
        .reset_index(drop=True)
    )

    summary = pd.DataFrame([
        ["Total tenants analyzed", full["tenant_id"].nunique()],
        ["Total pipelines/dags analyzed", full["dag_id"].nunique()],
        ["Total (tenant, dag) combinations", len(full)],
        ["Missing schedules (need to add)", len(missing)],
        ["  - Always-on (no flag needed)", (missing["dependency_type"] == "Always-on (no flag)").sum()],
        ["  - Flag-gated (open flag exists)", (missing["dependency_type"] == "Flag-gated").sum()],
        ["Extra schedules to review (flag closed)", len(to_remove)],
        ["Unmapped dag rows (no flag / unknown flag)", (full["required_by_flags"] == "UNMAPPED").sum()],
        ["Already correctly scheduled", (full["action"] == "OK - already scheduled").sum()],
        ["Already correctly NOT scheduled", (full["action"] == "OK - correctly not scheduled").sum()],
    ], columns=["Metric", "Value"])

    return missing, to_remove, unmapped, summary


# ---------------------------------------------------------------------------
# 3. EXCEL OUTPUT
# ---------------------------------------------------------------------------

def write_excel(output_path, summary, missing, to_remove, unmapped, full):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        missing.to_excel(writer, sheet_name="Missing_Schedules", index=False)
        to_remove.to_excel(writer, sheet_name="Extra_Schedules_Review", index=False)
        unmapped.to_excel(writer, sheet_name="Unmapped_Features", index=False)
        full.to_excel(writer, sheet_name="Full_Detail", index=False)

    wb = openpyxl.load_workbook(output_path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col, max_row = ws.max_column, ws.max_row

        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        if max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        widths = {}
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 2000)):
            for cell in row:
                if cell.row > 1:
                    cell.font = body_font
                length = len(str(cell.value)) if cell.value is not None else 0
                widths[cell.column] = max(widths.get(cell.column, 0), length)
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w + 2, 10), 55)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# 4. MAIN
# ---------------------------------------------------------------------------

def main():
    dag_tags, dag_tables, tenant_scheduled, all_tenants = load_pipelines_json(PIPELINES_JSON_PATH)
    tenant_open_flags, all_known_flags = load_tenant_flags(FEATURES_FLAGS_PATH)

    full = build_report(dag_tags, dag_tables, tenant_open_flags,
                         all_known_flags, tenant_scheduled, all_tenants)
    missing, to_remove, unmapped, summary = slice_report(full)

    write_excel(OUTPUT_PATH, summary, missing, to_remove, unmapped, full)
    print(f"Report written to {OUTPUT_PATH}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()