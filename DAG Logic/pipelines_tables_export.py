"""
Pipelines -> Tables Export
===========================
Reads pipelines.json and writes an Excel file with one row per
(tenant_id, dag_id, table) - i.e. dags with multiple tables get exploded
into separate rows, one per table.

Usage:
  Edit the paths in the CONFIG section below and run:
      python pipelines_tables_export.py
"""

import json

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG - edit these paths before running
# ---------------------------------------------------------------------------

INPUT_PATH = "C:\\Users\\suraj\\OneDrive\\Desktop\\new 6.json"
OUTPUT_PATH = "C:\\Users\\suraj\\OneDrive\\Desktop\\pipelines_tables.xlsx"


def load_rows(path: str):
    with open(path) as f:
        payload = json.load(f)

    records = payload["results"] if isinstance(payload, dict) and "results" in payload else payload

    rows = []
    for rec in records:
        tenant_id = str(rec.get("tenant_id", "")).strip()
        dag_id = str(rec.get("dag_id", "")).strip()
        tables = rec.get("tables") or []
        feature_flags = ", ".join(rec.get("feature_flags") or [])

        if tables:
            for table in tables:
                rows.append({"tenant_id": tenant_id, "dag_id": dag_id, "tables": table, "feature_flags": feature_flags})
        else:
            rows.append({"tenant_id": tenant_id, "dag_id": dag_id, "tables": "", "feature_flags": feature_flags})

    return pd.DataFrame(rows)


def write_excel(output_path, df: pd.DataFrame):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Pipelines_Tables", index=False)

    wb = openpyxl.load_workbook(output_path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial")

    ws = wb["Pipelines_Tables"]
    max_col, max_row = ws.max_column, ws.max_row

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    widths = {}
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, 2000)):
        for cell in row:
            if cell.row > 1:
                cell.font = body_font
            length = len(str(cell.value)) if cell.value is not None else 0
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w + 2, 10), 55)

    wb.save(output_path)


def main():
    df = load_rows(INPUT_PATH)
    write_excel(OUTPUT_PATH, df)
    print(f"Report written to {OUTPUT_PATH}")
    print(f"Rows: {len(df)}")


if __name__ == "__main__":
    main()
